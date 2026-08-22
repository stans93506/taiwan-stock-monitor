"""
主動ETF持股追蹤模組
每日執行一次，抓取 ezmoney ETF 持股 Excel，存 JSON 快照，比對前日變動。
"""
import io, json, math, os, re, requests, urllib.parse
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl

DATA_DIR = Path(__file__).parent / "etf_data"
DATA_DIR.mkdir(exist_ok=True)

# ETF 設定：code → {"name", "source", "param"}
# source: "ezmoney"    → param = ezmoney fund_code (str)
#         "fhtrust"    → param = fhtrust detail_id (ETFxx)
#         "capitalfund"→ param = fund_id (int)
#         "cathayfund" → param = internal fund_code (str, e.g. "EA")
#         "nomura"     → param = ETF stock code (str, e.g. "00980A")
#         "allianzgi"  → param = fund ID (str, e.g. "E0002")
#         "ctbcfund"   → param = ETF stock code (str, e.g. "00995A")
#         "megafund"   → param = page id (int, e.g. 23)
#         "tsit"       → param = ETF stock code (str, e.g. "00987A")
#         "fsitc"      → param = FundDetail page ID (str, e.g. "182")
#         "jpmorgan"   → param = CUSIP/ISIN code (str, e.g. "TW00000401A1")
ETF_CONFIG = {
    "00981A": {"name": "主動統一台股增長",    "source": "ezmoney",    "param": "49YTW"},
    "00403A": {"name": "主動統一升級50",      "source": "ezmoney",    "param": "63YTW"},
    "00991A": {"name": "主動復華未來50 ETF",  "source": "fhtrust",    "param": "ETF23"},
    "00982A": {"name": "主動群益台灣強棒",    "source": "capitalfund","param": 399},
    "00992A": {"name": "主動群益科技創新",    "source": "capitalfund","param": 500},
    "00400A": {"name": "主動國泰動能高息",    "source": "cathayfund", "param": "EA"},
    "00980A": {"name": "主動野村臺灣優選",    "source": "nomura",     "param": "00980A"},
    "00985A": {"name": "主動野村台灣50",      "source": "nomura",     "param": "00985A"},
    "00993A": {"name": "主動安聯台灣",        "source": "allianzgi",  "param": "E0002"},
    "00984A": {"name": "主動安聯台灣高息",    "source": "allianzgi",  "param": "E0001"},
    "00995A": {"name": "主動中信台灣卓越",    "source": "ctbcfund",   "param": "00995A"},
    "00996A": {"name": "主動兆豐台灣要收",    "source": "megafund",   "param": 23},
    "00987A": {"name": "主動台新優勢成長",    "source": "tsit",       "param": "00987A"},
    "00994A": {"name": "主動第一金台股優選",  "source": "fsitc",      "param": "182"},
    "00401A": {"name": "主動摩根台灣鑫收",   "source": "jpmorgan",   "param": "TW00000401A1"},
}

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Referer": "https://www.ezmoney.com.tw/ETF/Fund/Info/",
}
BASE_URL = "https://www.ezmoney.com.tw"


# ── 資料抓取 ──────────────────────────────────────────────────────────

def _make_session(etf_code: str) -> requests.Session:
    s = requests.Session()
    s.get(f"{BASE_URL}/ETF/Fund/Info/{etf_code}", headers=HEADERS, timeout=15)
    return s


def fetch_holdings(etf_code: str, fund_code: str) -> dict | None:
    """
    下載 ezmoney Excel，解析成 dict：
    {date, nav_per_unit, nav_total, units, holdings: [{code,name,shares,weight}]}
    """
    s = _make_session(etf_code)
    try:
        r = s.get(f"{BASE_URL}/ETF/Fund/AssetExcelNPOI?fundCode={fund_code}",
                  headers=HEADERS, timeout=20)
        if r.status_code != 200 or len(r.content) < 1000:
            print(f"  [ETF] {etf_code} 下載失敗 HTTP={r.status_code}")
            return None
    except Exception as e:
        print(f"  [ETF] {etf_code} 連線失敗：{e}")
        return None

    try:
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        ws = wb.worksheets[0]
        rows = [row for row in ws.iter_rows(values_only=True)
                if any(c is not None for c in row)]
    except Exception as e:
        print(f"  [ETF] {etf_code} Excel 解析失敗：{e}")
        return None

    # 解析日期
    date_str = ""
    for row in rows:
        if row[0] and "資料日期" in str(row[0]):
            m = re.search(r'(\d{3}/\d{2}/\d{2})', str(row[0]))
            if m:
                date_str = m.group(1)
            break

    # 解析 NAV
    nav_total = nav_per_unit = units = None
    for row in rows:
        v0 = str(row[0] or "")
        v1 = str(row[1] or "")
        if "淨資產" in v0 and "NTD" in v1:
            nav_total = float(v1.replace("NTD", "").replace(",", "").strip())
        elif "流通在外" in v0:
            units = float(v1.replace(",", "").strip())
        elif "每單位淨值" in v0 and "NTD" in v1:
            nav_per_unit = float(v1.replace("NTD", "").replace(",", "").strip())

    # 解析持股清單（找 header 列）
    holdings = []
    in_stocks = False
    for row in rows:
        v0 = str(row[0] or "")
        if "股票代號" in v0:
            in_stocks = True
            continue
        if not in_stocks:
            continue
        code = str(row[0] or "").strip()
        if not code or not code[:1].isdigit():
            continue
        try:
            shares = int(str(row[2] or "0").replace(",", "").strip())
            weight = float(str(row[3] or "0").replace("%", "").replace(",", "").strip())
            holdings.append({
                "code":   code,
                "name":   str(row[1] or "").strip(),
                "shares": shares,
                "weight": weight,
            })
        except Exception:
            continue

    return {
        "etf_code":     etf_code,
        "etf_name":     ETF_CONFIG.get(etf_code, {}).get("name", etf_code),
        "date":         date_str,
        "nav_per_unit": nav_per_unit,
        "nav_total":    nav_total,
        "units":        units,
        "holdings":     holdings,
    }


def fetch_holdings_fhtrust(etf_code: str, detail_id: str) -> dict | None:
    """
    從復華基金官網 fhtrust.com.tw 下載 ETF 持股 Excel 並解析。
    日期格式西元 YYYY/MM/DD → 民國 RRR/MM/DD。
    """
    base = "https://www.fhtrust.com.tw"
    h = {
        "User-Agent": HEADERS["User-Agent"],
        "Referer": f"{base}/ETF/etf_detail/{detail_id}",
    }

    # Step1：從今天往回找最新有效 Excel（HTML 寫死舊日期，需直接試）
    from datetime import timedelta
    r = None
    excel_url = None
    probe_date = datetime.today()
    for _ in range(10):  # 最多往回找 10 天
        candidate = f"{base}/api/assetsExcel/{detail_id}/{probe_date.strftime('%Y%m%d')}"
        try:
            resp = requests.get(candidate, headers=h, timeout=10, verify=False)
            if resp.status_code == 200 and len(resp.content) > 500:
                r = resp
                excel_url = candidate
                break
        except Exception:
            pass
        probe_date -= timedelta(days=1)

    if r is None:
        print(f"\n  [fhtrust] {etf_code} 找不到有效 Excel（往回找 10 天皆失敗）")
        return None

    # Step3：解析 Excel
    try:
        wb = openpyxl.load_workbook(io.BytesIO(r.content), data_only=True)
        ws = wb.active
        rows = [row for row in ws.iter_rows(values_only=True)
                if any(c is not None for c in row)]
    except Exception as e:
        print(f"\n  [fhtrust] {etf_code} Excel 解析失敗：{e}")
        return None

    # 解析日期（西元 YYYY/MM/DD → 民國 RRR/MM/DD）
    date_str = ""
    for row in rows:
        v0 = str(row[0] or "")
        if "日期" in v0:
            m = re.search(r'(\d{4})/(\d{2}/\d{2})', v0)
            if m:
                roc = int(m.group(1)) - 1911
                date_str = f"{roc}/{m.group(2)}"
            break

    # 解析 NAV（每個標籤在 col0，值在下一列的 col0）
    nav_total = nav_per_unit = units = None
    for i, row in enumerate(rows):
        v0 = str(row[0] or "")
        nxt = str(rows[i + 1][0] or "") if i + 1 < len(rows) else ""
        if "基金資產淨值" in v0:
            try:
                nav_total = float(nxt.replace(",", "").strip())
            except Exception:
                pass
        elif "流通單位數" in v0:
            try:
                units = float(nxt.replace(",", "").strip())
            except Exception:
                pass
        elif "每單位淨值" in v0:
            try:
                nav_per_unit = float(nxt.replace(",", "").strip())
            except Exception:
                pass

    # 解析持股（"證券代號" 後開始；code=col0, name=col1, shares=col2, weight=col4）
    holdings = []
    in_stocks = False
    for row in rows:
        v0 = str(row[0] or "")
        if "證券代號" in v0:
            in_stocks = True
            continue
        if not in_stocks:
            continue
        code = v0.strip()
        if not code or not code[:1].isdigit():
            continue
        try:
            shares = int(str(row[2] or "0").replace(",", "").strip())
            weight = float(str(row[4] or "0").replace("%", "").replace(",", "").strip())
            holdings.append({
                "code":   code,
                "name":   str(row[1] or "").strip(),
                "shares": shares,
                "weight": weight,
            })
        except Exception:
            continue

    return {
        "etf_code":     etf_code,
        "etf_name":     ETF_CONFIG.get(etf_code, {}).get("name", etf_code),
        "date":         date_str,
        "nav_per_unit": nav_per_unit,
        "nav_total":    nav_total,
        "units":        units,
        "holdings":     holdings,
    }


def fetch_holdings_capitalfund(etf_code: str, fund_id: int) -> dict | None:
    """
    從群益投信官網 capitalfund.com.tw 取得 ETF 持股（PCF 申購買回清單）
    POST /CFWeb/api/etf/buyback → data.stocks
    日期格式西元 YYYY-MM-DD → 民國 RRR/MM/DD
    """
    base = "https://www.capitalfund.com.tw"
    h = {
        "User-Agent": HEADERS["User-Agent"],
        "Referer": f"{base}/etf/product/detail/{fund_id}/portfolio",
        "Accept": "application/json",
        "Content-Type": "application/json",
    }
    try:
        r = requests.post(f"{base}/CFWeb/api/etf/buyback",
                          json={"fundId": fund_id},
                          headers=h, timeout=20, verify=False)
        if r.status_code != 200:
            print(f"\n  [capital] {etf_code} API 失敗 HTTP={r.status_code}")
            return None
        resp = r.json()
        if resp.get("code") != 200:
            print(f"\n  [capital] {etf_code} API 錯誤: {resp.get('message','')}")
            return None
    except Exception as e:
        print(f"\n  [capital] {etf_code} 連線失敗：{e}")
        return None

    pcf    = resp["data"].get("pcf", {})
    stocks = resp["data"].get("stocks", [])

    # 解析日期（西元 "2026-07-24" → 民國 "115/07/24"）
    date_str = ""
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})", str(pcf.get("date2", "")))
    if m:
        date_str = f"{int(m.group(1))-1911}/{m.group(2)}/{m.group(3)}"

    # 解析持股
    holdings = []
    for item in stocks:
        code = str(item.get("stocNo", "")).strip()
        if not code or not code[:1].isdigit():
            continue
        try:
            shares = int(float(str(item.get("share", 0)).replace(",", "")))
            weight = float(str(item.get("weight", 0)))
            holdings.append({
                "code":   code,
                "name":   str(item.get("stocName", "")).strip(),
                "shares": shares,
                "weight": weight,
            })
        except Exception:
            continue

    return {
        "etf_code":     etf_code,
        "etf_name":     ETF_CONFIG.get(etf_code, {}).get("name", etf_code),
        "date":         date_str,
        "nav_per_unit": pcf.get("pUnit"),
        "nav_total":    pcf.get("nav"),
        "units":        pcf.get("totUnit"),
        "holdings":     holdings,
    }


def fetch_holdings_cathayfund(etf_code: str, fund_code: str) -> dict | None:
    """
    從國泰投信 cwapi.cathaysite.com.tw 取得 ETF 持股。
    先呼叫 GetETFAssets 取最新資料日期，再呼叫 GetETFDetailStockList。
    日期格式西元 YYYY/MM/DD → 民國 RRR/MM/DD。
    """
    base = "https://cwapi.cathaysite.com.tw/"
    h = {
        "User-Agent": HEADERS["User-Agent"],
        "Referer": "https://www.cathaysite.com.tw/ETF/detail/",
        "Origin": "https://www.cathaysite.com.tw",
        "Accept": "application/json",
    }

    # Step1：取最新資料日期
    try:
        r0 = requests.get(f"{base}api/ETF/GetETFAssets", params={"FundCode": fund_code},
                          headers=h, timeout=15, verify=False)
        assets = r0.json()
        if assets.get("returnCode") != "2000":
            print(f"\n  [cathayfund] {etf_code} GetETFAssets 失敗: {assets.get('returnMessage','')}")
            return None
        a = assets["result"]
        search_date = a.get("preDate", "")   # "2026/07/24"
        nav_total    = float(str(a.get("fundNav", "0")).replace(",", "") or 0)
        nav_per_unit = float(str(a.get("fundPerNav", "0")).replace(",", "") or 0)
        units        = float(str(a.get("fundOutstandingShares", "0")).replace(",", "") or 0)
    except Exception as e:
        print(f"\n  [cathayfund] {etf_code} GetETFAssets 失敗：{e}")
        return None

    # Step2：取持股清單
    try:
        r = requests.get(f"{base}api/ETF/GetETFDetailStockList",
                         params={"FundCode": fund_code, "SearchDate": search_date},
                         headers=h, timeout=20, verify=False)
        resp = r.json()
        if resp.get("returnCode") != "2000":
            print(f"\n  [cathayfund] {etf_code} GetETFDetailStockList 失敗: {resp.get('returnMessage','')}")
            return None
        stocks = resp.get("result") or []
    except Exception as e:
        print(f"\n  [cathayfund] {etf_code} 連線失敗：{e}")
        return None

    # 解析日期（西元 YYYY/MM/DD → 民國 RRR/MM/DD）
    date_str = ""
    m = re.match(r"(\d{4})/(\d{2}/\d{2})", search_date)
    if m:
        date_str = f"{int(m.group(1))-1911}/{m.group(2)}"

    holdings = []
    for item in stocks:
        code = str(item.get("stockCode", "")).strip()
        if not code or not code[:1].isdigit():
            continue
        try:
            shares = int(str(item.get("volumn", "0")).replace(",", ""))
            weight = float(str(item.get("weights", "0")).replace(",", ""))
            holdings.append({
                "code":   code,
                "name":   str(item.get("stockName", "")).strip(),
                "shares": shares,
                "weight": weight,
            })
        except Exception:
            continue

    return {
        "etf_code":     etf_code,
        "etf_name":     ETF_CONFIG.get(etf_code, {}).get("name", etf_code),
        "date":         date_str,
        "nav_per_unit": nav_per_unit,
        "nav_total":    nav_total,
        "units":        units,
        "holdings":     holdings,
    }


def fetch_holdings_allianzgi(etf_code: str, fund_id: str) -> dict | None:
    """
    從安聯投信 etf.allianzgi.com.tw 取得 ETF 持股。
    先 GET GetAntiForgeryToken 取得 X-XSRF-TOKEN cookie，
    再 POST Fund/GetFundAssets {FundID: fund_id}。
    Row format: [序號, code, name, shares_str, weight%_str]
    日期格式西元 YYYY/MM/DD → 民國 RRR/MM/DD。
    """
    base = "https://etf.allianzgi.com.tw/webapi/api/"
    h = {
        "User-Agent": HEADERS["User-Agent"],
        "Referer": f"https://etf.allianzgi.com.tw/etf-info/{fund_id}?tab=4",
        "Origin": "https://etf.allianzgi.com.tw",
        "Accept": "application/json, text/plain, */*",
    }
    s = requests.Session()
    s.verify = False

    try:
        s.get(f"{base}AntiForgery/GetAntiForgeryToken", headers=h, timeout=15)
        xsrf = s.cookies.get("X-XSRF-TOKEN", "")
        if not xsrf:
            print(f"\n  [allianz] {etf_code} 無法取得 XSRF token")
            return None
    except Exception as e:
        print(f"\n  [allianz] {etf_code} token 取得失敗：{e}")
        return None

    post_h = {**h, "Content-Type": "application/json", "X-XSRF-TOKEN": xsrf}
    try:
        r = s.post(f"{base}Fund/GetFundAssets",
                   json={"FundID": fund_id},
                   headers=post_h, timeout=20)
        if r.status_code != 200:
            print(f"\n  [allianz] {etf_code} HTTP {r.status_code}")
            return None
        data = r.json()
        d = data["Entries"]["Data"]
        fa = d.get("FundAsset") or {}
        tables = d.get("Table") or []
    except Exception as e:
        print(f"\n  [allianz] {etf_code} 連線失敗：{e}")
        return None

    nav_total    = float(str(fa.get("Aum",   "0")).replace(",", "") or 0)
    nav_per_unit = float(str(fa.get("Nav",   "0")).replace(",", "") or 0)
    units        = float(str(fa.get("Units", "0")).replace(",", "") or 0)
    raw_date     = fa.get("NavDate", "")  # "2026/07/24"

    date_str = ""
    m = re.match(r"(\d{4})/(\d{2}/\d{2})", raw_date)
    if m:
        date_str = f"{int(m.group(1))-1911}/{m.group(2)}"

    holdings = []
    for tbl in tables:
        if "股票" not in (tbl.get("TableTitle") or ""):
            continue
        for row in (tbl.get("Rows") or []):
            if len(row) < 5:
                continue
            code = str(row[1]).strip()
            if not code or not code[:1].isdigit():
                continue
            try:
                shares = int(str(row[3]).replace(",", ""))
                weight = float(str(row[4]).replace("%", "").replace(",", ""))
                holdings.append({
                    "code":   code,
                    "name":   str(row[2]).strip(),
                    "shares": shares,
                    "weight": weight,
                })
            except Exception:
                continue
        break

    return {
        "etf_code":     etf_code,
        "etf_name":     ETF_CONFIG.get(etf_code, {}).get("name", etf_code),
        "date":         date_str,
        "nav_per_unit": nav_per_unit,
        "nav_total":    nav_total,
        "units":        units,
        "holdings":     holdings,
    }


def fetch_holdings_nomura(etf_code: str, fund_id: str) -> dict | None:
    """
    從野村投信 nomurafunds.com.tw 取得 ETF 持股。
    POST /API/ETFAPI/api/Fund/GetFundAssets，SearchDate=null 拿最新。
    回傳格式：Entries.Data.FundAsset + Entries.Data.Table[0].Rows。
    日期格式西元 YYYY/MM/DD → 民國 RRR/MM/DD。
    """
    base = "https://www.nomurafunds.com.tw/API/ETFAPI/api/"
    h = {
        "User-Agent": HEADERS["User-Agent"],
        "Referer": f"https://www.nomurafunds.com.tw/ETFWEB/product-description?fundNo={fund_id}&tab=Shareholding",
        "Origin": "https://www.nomurafunds.com.tw",
        "Content-Type": "application/json",
        "Accept": "application/json",
    }
    try:
        r = requests.post(f"{base}Fund/GetFundAssets",
                          json={"FundID": fund_id, "SearchDate": None},
                          headers=h, timeout=20, verify=False)
        if r.status_code != 200:
            print(f"\n  [nomura] {etf_code} HTTP {r.status_code}")
            return None
        data = r.json()
        entries = (data.get("Entries") or {})
        d = (entries.get("Data") or {})
        fa = d.get("FundAsset") or {}
        tables = d.get("Table") or []
    except Exception as e:
        print(f"\n  [nomura] {etf_code} 連線失敗：{e}")
        return None

    # NAV
    nav_total    = float(str(fa.get("Aum", "0")).replace(",", "") or 0)
    nav_per_unit = float(str(fa.get("Nav", "0")).replace(",", "") or 0)
    units        = float(str(fa.get("Units", "0")).replace(",", "") or 0)
    raw_date     = fa.get("NavDate", "")  # "2026/07/24"

    # 日期轉民國
    date_str = ""
    m = re.match(r"(\d{4})/(\d{2}/\d{2})", raw_date)
    if m:
        date_str = f"{int(m.group(1))-1911}/{m.group(2)}"

    # 持股：找第一個 TableTitle == "股票" 的表格
    holdings = []
    for tbl in tables:
        if tbl.get("TableTitle") != "股票":
            continue
        for row in (tbl.get("Rows") or []):
            if len(row) < 4:
                continue
            code = str(row[0]).strip()
            if not code or not code[:1].isdigit():
                continue
            try:
                shares = int(str(row[2]).replace(",", ""))
                weight = float(str(row[3]).replace(",", ""))
                holdings.append({
                    "code":   code,
                    "name":   str(row[1]).strip(),
                    "shares": shares,
                    "weight": weight,
                })
            except Exception:
                continue
        break  # 只取第一個股票表

    return {
        "etf_code":     etf_code,
        "etf_name":     ETF_CONFIG.get(etf_code, {}).get("name", etf_code),
        "date":         date_str,
        "nav_per_unit": nav_per_unit,
        "nav_total":    nav_total,
        "units":        units,
        "holdings":     holdings,
    }


def fetch_holdings_jpmorgan(etf_code: str, cusip: str) -> dict | None:
    """
    從摩根資產管理 am.jpmorgan.com 取得 ETF 持股。
    GET FundsMarketingHandler/product-data?cusip={cusip}&country=tw&role=twetf...
    持股在 fundData.holdings.pcfEquityHoldings.data，
    NAV 在 fundData.shareClass.nav / netGrossAssetClass / sharesOutstanding。
    Fields: securityTicker=code, securityDescription=name, shares, marketValuePercent=weight%。
    """
    url = "https://am.jpmorgan.com/FundsMarketingHandler/product-data"
    params = {
        "cusip": cusip,
        "country": "tw",
        "role": "twetf",
        "language": "zh",
        "userLoggedIn": "false",
    }
    h = {
        "User-Agent": HEADERS["User-Agent"],
        "Accept": "application/json",
        "Referer": "https://am.jpmorgan.com/tw/zh/asset-management/twetf/products/",
    }

    try:
        r = requests.get(url, params=params, headers=h, timeout=30, verify=False)
        if r.status_code != 200:
            print(f"\n  [jpmorgan] {etf_code} HTTP {r.status_code}")
            return None
        fd = r.json()["fundData"]
    except Exception as e:
        print(f"\n  [jpmorgan] {etf_code} 連線失敗：{e}")
        return None

    # NAV
    sc = fd.get("shareClass") or {}
    nav_obj = sc.get("nav") or {}
    nav_per_unit = float(nav_obj.get("price") or 0)
    nav_total    = float(str(sc.get("netGrossAssetClass") or 0))
    units        = float(str(sc.get("sharesOutstanding") or 0))

    raw_date = nav_obj.get("date") or fd.get("stringValueWrapper", {}).get("holdingLatestDate") or ""
    date_str = ""
    md = re.match(r"(\d{4})-(\d{2})-(\d{2})", raw_date)
    if md:
        date_str = f"{int(md.group(1))-1911}/{md.group(2)}/{md.group(3)}"

    # 持股
    eq = (fd.get("holdings") or {}).get("pcfEquityHoldings") or {}
    raw_holdings = eq.get("data") or []

    holdings = []
    for row in raw_holdings:
        code = str(row.get("securityTicker") or "").strip()
        if not code or not code[:1].isdigit():
            continue
        try:
            shares = int(row.get("shares") or 0)
            weight = float(row.get("marketValuePercent") or 0)
            holdings.append({
                "code":   code,
                "name":   str(row.get("securityDescription") or "").strip(),
                "shares": shares,
                "weight": weight,
            })
        except Exception:
            continue

    return {
        "etf_code":     etf_code,
        "etf_name":     ETF_CONFIG.get(etf_code, {}).get("name", etf_code),
        "date":         date_str,
        "nav_per_unit": nav_per_unit,
        "nav_total":    nav_total,
        "units":        units,
        "holdings":     holdings,
    }


def fetch_holdings_fsitc(etf_code: str, fund_id: str) -> dict | None:
    """
    從第一金投信 fsitc.com.tw 取得 ETF 持股。
    先 GET FundDetail 頁面取得 session cookie，
    再 POST WebAPI.aspx/Get_hd (空 date) → 最新持股，
    POST WebAPI.aspx/Get_BuySellA → 最新 NAV。
    Row fields: A=code, B=name, C=weight%, D=shares, group=1 為股票。
    NAV rows: A=標籤, B=數值，匹配「基金淨資產」/「每受益權」/「已發行受益權」。
    """
    base = "https://www.fsitc.com.tw/"
    h_page = {"User-Agent": HEADERS["User-Agent"]}
    h_api = {
        "User-Agent": HEADERS["User-Agent"],
        "Content-Type": "application/json; charset=utf-8",
        "Accept": "application/json, text/javascript, */*; q=0.01",
        "Referer": f"{base}FundDetail.aspx?ID={fund_id}",
        "X-Requested-With": "XMLHttpRequest",
    }

    s = requests.Session()
    s.verify = False
    try:
        s.get(f"{base}FundDetail.aspx?ID={fund_id}", headers=h_page, timeout=20)
    except Exception as e:
        print(f"\n  [fsitc] {etf_code} 頁面讀取失敗：{e}")
        return None

    def _call(endpoint):
        payload = json.dumps({"pStrFundID": fund_id, "pStrDate": ""})
        r = s.post(f"{base}WebAPI.aspx/{endpoint}", data=payload, headers=h_api, timeout=15)
        raw = r.json().get("d", "")
        return json.loads(raw) if isinstance(raw, str) and raw else (raw or [])

    try:
        hd_rows = _call("Get_hd")
        nav_rows = _call("Get_BuySellA")
    except Exception as e:
        print(f"\n  [fsitc] {etf_code} API 呼叫失敗：{e}")
        return None

    # 日期 "2026-07-24" → "115/07/24"
    raw_date = hd_rows[0].get("sdate", "") if hd_rows else ""
    date_str = ""
    md = re.match(r"(\d{4})-(\d{2})-(\d{2})", raw_date)
    if md:
        date_str = f"{int(md.group(1))-1911}/{md.group(2)}/{md.group(3)}"

    # NAV
    def _nav_val(label_kw):
        for row in nav_rows:
            if label_kw in (row.get("A") or ""):
                return str(row.get("B", "0")).replace(",", "").replace("TWD", "").strip()
        return "0"

    nav_total    = float(_nav_val("基金淨資產價值") or 0)
    nav_per_unit = float(_nav_val("每受益權單位淨資產價值") or 0)
    units        = float(_nav_val("已發行受益權單位總數") or 0)

    # 持股：只取 group="1"（股票）
    holdings = []
    for row in hd_rows:
        if row.get("group") != "1":
            continue
        code = str(row.get("A", "")).strip()
        if not code or not code[:1].isdigit():
            continue
        try:
            shares = int(str(row.get("D", "0")).replace(",", ""))
            weight = float(str(row.get("C", "0")).replace(",", ""))
            holdings.append({
                "code":   code,
                "name":   str(row.get("B", "")).strip(),
                "shares": shares,
                "weight": weight,
            })
        except Exception:
            continue

    return {
        "etf_code":     etf_code,
        "etf_name":     ETF_CONFIG.get(etf_code, {}).get("name", etf_code),
        "date":         date_str,
        "nav_per_unit": nav_per_unit,
        "nav_total":    nav_total,
        "units":        units,
        "holdings":     holdings,
    }


def fetch_holdings_ctbcfund(etf_code: str, etf_id: str) -> dict | None:
    """
    從中信投信 ctbcinvestments.com.tw 取得 ETF 持股。
    步驟：
    1. POST AuthToken → token
    2. POST ETFList (空 body) → 找 ETF_ID 匹配 → 取 FID 和 NAV_DT
    3. POST ETFHoldingWeight {FID, StartDate} → 持股明細
    """
    base = "https://www.ctbcinvestments.com.tw/API/"
    h = {
        "User-Agent": HEADERS["User-Agent"],
        "Referer": "https://www.ctbcinvestments.com/",
        "Origin": "https://www.ctbcinvestments.com",
        "Accept": "application/json",
        "Content-Type": "application/json",
    }

    try:
        r0 = requests.post(base + "home/AuthToken?token=www.ctbcinvestments.com",
                           headers=h, timeout=15, verify=False)
        token = r0.json()["Data"]["token"]
        enc = urllib.parse.quote(token, safe="")
    except Exception as e:
        print(f"\n  [ctbc] {etf_code} token 失敗：{e}")
        return None

    try:
        r1 = requests.post(base + f"etf/ETFList?token={enc}",
                           headers=h, json={}, timeout=15, verify=False)
        etf_list = r1.json().get("Data", {}).get("Data", [])
        item = next((x for x in etf_list if x.get("ETF_ID") == etf_id), None)
        if not item:
            print(f"\n  [ctbc] {etf_code} 找不到 ETF_ID={etf_id}")
            return None
        fid = item["FID"]
        nav_dt = item.get("NAV_DT", "")[:10].replace("-", "/")
    except Exception as e:
        print(f"\n  [ctbc] {etf_code} ETFList 失敗：{e}")
        return None

    try:
        r2 = requests.post(base + f"etf/ETFHoldingWeight?token={enc}",
                           headers=h, json={"FID": fid, "StartDate": nav_dt},
                           timeout=20, verify=False)
        data = r2.json().get("Data", {})
    except Exception as e:
        print(f"\n  [ctbc] {etf_code} ETFHoldingWeight 失敗：{e}")
        return None

    fa_list = data.get("FundAssets", [])
    fa = fa_list[0] if fa_list else {}
    nav_total    = float(str(fa.get("基金淨資產",         "0")).replace(",", "") or 0)
    units        = float(str(fa.get("基金在外流通單位數",  "0")).replace(",", "") or 0)
    nav_per_unit = float(str(fa.get("基金每單位淨值",      "0")).replace(",", "") or 0)
    raw_date     = fa.get("資料日期", nav_dt)

    date_str = ""
    m = re.match(r"(\d{4})/(\d{2}/\d{2})", raw_date)
    if m:
        date_str = f"{int(m.group(1))-1911}/{m.group(2)}"

    holdings = []
    for detail in data.get("FundAssetsDetail", []):
        if detail.get("Code") != "STOCK":
            continue
        for row in (detail.get("Data") or []):
            code = str(row.get("code_", "")).strip()
            if not code or not code[:1].isdigit():
                continue
            try:
                shares = int(float(str(row.get("qty_",     "0")).replace(",", "")))
                weight = float(str(row.get("weights_", "0")).replace("%", "").replace(",", ""))
                holdings.append({
                    "code":   code,
                    "name":   str(row.get("name_", "")).strip(),
                    "shares": shares,
                    "weight": weight,
                })
            except Exception:
                continue
        break

    return {
        "etf_code":     etf_code,
        "etf_name":     ETF_CONFIG.get(etf_code, {}).get("name", etf_code),
        "date":         date_str,
        "nav_per_unit": nav_per_unit,
        "nav_total":    nav_total,
        "units":        units,
        "holdings":     holdings,
    }


def fetch_holdings_megafund(etf_code: str, page_id: int) -> dict | None:
    """
    從兆豐投信 megafunds.com.tw 取得 ETF 持股。
    GET etf_product.aspx?id={page_id} 解析 HTML。
    """
    url = f"https://www.megafunds.com.tw/MEGA/etf/etf_product.aspx?id={page_id}"
    h = {"User-Agent": HEADERS["User-Agent"]}

    try:
        r = requests.get(url, headers=h, timeout=20, verify=False)
        if r.status_code != 200:
            print(f"\n  [megafund] {etf_code} HTTP {r.status_code}")
            return None
        html = r.text
    except Exception as e:
        print(f"\n  [megafund] {etf_code} 連線失敗：{e}")
        return None

    # 日期
    m_date = re.search(r'資料來源：兆豐投信，(\d{4}/\d{2}/\d{2})', html)
    raw_date = m_date.group(1) if m_date else ""
    date_str = ""
    md = re.match(r"(\d{4})/(\d{2}/\d{2})", raw_date)
    if md:
        date_str = f"{int(md.group(1))-1911}/{md.group(2)}"

    # NAV
    m_nav = re.search(r'每單位淨值</div>\s*<div[^>]*class="si-amount"[^>]*>\s*([\d.]+)', html)
    nav_per_unit = float(m_nav.group(1)) if m_nav else 0.0

    m_units = re.search(r'在外流通單位數</div>\s*<div[^>]*class="si-amount"[^>]*>\s*([\d,]+)', html)
    units = float(str(m_units.group(1)).replace(",", "")) if m_units else 0.0
    nav_total = round(nav_per_unit * units)

    # 持股 (4 columns per row)
    holdings_raw = re.findall(
        r'<div class="fund-info content-list-1">\s*'
        r'<div class="fund-content">(\d{4})</div>\s*'
        r'<div class="fund-content">([^<]+)</div>\s*'
        r'<div class="fund-content txt-right">([^<]+)</div>\s*'
        r'<div class="fund-content txt-right">([^<]+)</div>',
        html
    )

    holdings = []
    for code, name, shares_s, weight_s in holdings_raw:
        try:
            shares = int(shares_s.strip().replace(",", ""))
            weight = float(weight_s.strip().replace("%", "").replace(",", ""))
            holdings.append({"code": code, "name": name.strip(), "shares": shares, "weight": weight})
        except Exception:
            continue

    return {
        "etf_code":     etf_code,
        "etf_name":     ETF_CONFIG.get(etf_code, {}).get("name", etf_code),
        "date":         date_str,
        "nav_per_unit": nav_per_unit,
        "nav_total":    float(nav_total),
        "units":        units,
        "holdings":     holdings,
    }


def fetch_holdings_tsit(etf_code: str, fund_code: str) -> dict | None:
    """
    從台新投信 tsit.com.tw 取得 ETF 持股。
    GET /ETF/Home/ETFSeriesDetail/{fund_code} 解析 HTML。
    """
    url = f"https://www.tsit.com.tw/ETF/Home/ETFSeriesDetail/{fund_code}"
    h = {"User-Agent": HEADERS["User-Agent"]}

    try:
        r = requests.get(url, headers=h, timeout=20, verify=False)
        if r.status_code != 200:
            print(f"\n  [tsit] {etf_code} HTTP {r.status_code}")
            return None
        html = r.text
    except Exception as e:
        print(f"\n  [tsit] {etf_code} 連線失敗：{e}")
        return None

    # 日期：hidden input NAV_DATE = "2026/7/24 上午 12:00:00"
    m_date = re.search(r'NAV_DATE[^>]*value="(\d{4}/\d{1,2}/\d{1,2})', html)
    date_str = ""
    if m_date:
        md = re.match(r"(\d{4})/(\d{1,2})/(\d{1,2})", m_date.group(1))
        if md:
            yr = int(md.group(1)) - 1911
            date_str = f"{yr}/{int(md.group(2)):02d}/{int(md.group(3)):02d}"

    # NAV：「基金淨資產價值 TWD xxx」出現在 TWD amounts[0]；每單位淨值在[2]
    twd_amounts = re.findall(r'TWD\s*([\d,\.]+)', html)
    nav_total    = float(twd_amounts[0].replace(",", "")) if len(twd_amounts) > 0 else 0.0
    nav_per_unit = float(twd_amounts[2].replace(",", "")) if len(twd_amounts) > 2 else 0.0
    # 每單位淨值亦可用關鍵字定位作為備援
    if not nav_per_unit:
        mn = re.search(r'每單位淨值.*?(\d+\.\d+)', html, re.DOTALL)
        nav_per_unit = float(mn.group(1)) if mn else 0.0
    units = round(nav_total / nav_per_unit) if nav_per_unit else 0.0

    # 持股：<td>CODE TT</td><td>name</td><td>shares</td><td>weight%</td>
    holdings_raw = re.findall(
        r'<td>(\d{4})\s*TT</td>\s*<td>([^<]+)</td>\s*<td>([\d,]+)</td>\s*<td>([^<]+%)</td>',
        html
    )

    holdings = []
    for code, name, shares_s, weight_s in holdings_raw:
        try:
            shares = int(shares_s.strip().replace(",", ""))
            weight = float(weight_s.strip().replace("%", "").replace(",", ""))
            holdings.append({"code": code, "name": name.strip(), "shares": shares, "weight": weight})
        except Exception:
            continue

    return {
        "etf_code":     etf_code,
        "etf_name":     ETF_CONFIG.get(etf_code, {}).get("name", etf_code),
        "date":         date_str,
        "nav_per_unit": nav_per_unit,
        "nav_total":    nav_total,
        "units":        float(units),
        "holdings":     holdings,
    }


# ── 快照儲存 ──────────────────────────────────────────────────────────

def _snapshot_path(etf_code: str, date_str: str) -> Path:
    """民國日期 115/07/24 → etf_data/00981A_1150724.json"""
    safe = date_str.replace("/", "")
    return DATA_DIR / f"{etf_code}_{safe}.json"


def _today_snapshot_exists(etf_code: str) -> bool:
    """今日 ROC 日期的快照檔是否已存在（依檔名判斷，非 mtime）
    用檔名而非 mtime，避免「早上抓到昨天資料→存舊檔→以為今日已抓」的問題。
    """
    t = datetime.today()
    suffix = f"{t.year - 1911}{t.month:02d}{t.day:02d}"
    return (DATA_DIR / f"{etf_code}_{suffix}.json").exists()


def save_snapshot(data: dict) -> Path:
    path = _snapshot_path(data["etf_code"], data["date"])
    data.setdefault("fetch_time", (datetime.utcnow() + timedelta(hours=8)).strftime("%H:%M"))
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    return path


def load_latest_snapshots(etf_code: str, n: int = 2) -> list[dict]:
    """載入最新 n 份快照（依日期降冪）"""
    files = sorted(DATA_DIR.glob(f"{etf_code}_*.json"), reverse=True)
    result = []
    for p in files[:n]:
        try:
            with open(p, encoding="utf-8") as f:
                result.append(json.load(f))
        except Exception:
            pass
    return result


# ── 變動計算 ──────────────────────────────────────────────────────────

def compute_changes(today: dict, yesterday: dict | None) -> list[dict]:
    """
    比對今昨日持股，回傳合併列表：
    [{code, name, shares, weight, prev_shares, prev_weight,
      delta_shares, delta_weight, is_new, is_removed, value_twd}]
    """
    t_map = {h["code"]: h for h in today["holdings"]}
    y_map = {h["code"]: h for h in (yesterday["holdings"] if yesterday else [])}
    all_codes = sorted(set(t_map) | set(y_map))
    nav = today.get("nav_total") or 0

    rows = []
    for code in all_codes:
        t = t_map.get(code)
        y = y_map.get(code)
        cur_shares  = t["shares"]  if t else 0
        cur_weight  = t["weight"]  if t else 0.0
        prev_shares = y["shares"]  if y else 0
        prev_weight = y["weight"]  if y else 0.0
        rows.append({
            "code":         code,
            "name":         (t or y)["name"],
            "shares":       cur_shares,
            "weight":       cur_weight,
            "prev_shares":  prev_shares,
            "prev_weight":  prev_weight,
            "delta_shares": cur_shares - prev_shares,
            "delta_weight": round(cur_weight - prev_weight, 4),
            "is_new":       (t is not None and y is None),
            "is_removed":   (t is None and y is not None),
            "value_twd":    round(cur_weight / 100 * nav) if nav else 0,
        })
    return rows


# ── 當日漲跌幅 ───────────────────────────────────────────────────────

def fetch_stock_data(codes: list[str]) -> dict[str, dict]:
    """從 TWSE/OTC OpenAPI 抓取各股當日漲跌幅/成交量/已發行股數，回傳 {code: {pct, volume, issued}}"""
    if not codes:
        return {}
    result: dict[str, dict] = {}
    code_set = set(codes)

    def _pct(close_s, change_s) -> float | None:
        try:
            close  = float(str(close_s or "").replace(",", "").strip())
            change = float(str(change_s or "").replace("+", "").replace(",", "").strip())
            prev   = close - change
            return round(change / prev * 100, 2) if prev else 0.0
        except (ValueError, ZeroDivisionError, TypeError):
            return None

    def _int(s) -> int:
        try:
            return int(float(str(s or "0").replace(",", "")))
        except (ValueError, TypeError):
            return 0

    def _price(s) -> float | None:
        try:
            v = float(str(s or "").replace(",", "").strip())
            return v if v > 0 else None
        except (ValueError, TypeError):
            return None

    # TWSE 上市：漲跌幅 + 成交量 + 收盤價
    try:
        r = requests.get(
            "https://openapi.twse.com.tw/v1/exchangeReport/STOCK_DAY_ALL",
            headers={"Accept": "application/json"}, timeout=15, verify=False,
        )
        for item in (r.json() if r.ok else []):
            c = str(item.get("Code", "")).strip()
            if c not in code_set:
                continue
            result.setdefault(c, {})
            pct = _pct(item.get("ClosingPrice"), item.get("Change"))
            if pct is not None:
                result[c]["pct"] = pct
            result[c]["volume"] = _int(item.get("TradeVolume", 0))
            p = _price(item.get("ClosingPrice"))
            if p:
                result[c]["price"] = p
    except Exception as e:
        print(f"  ⚠ TWSE 漲跌幅/成交量查詢失敗：{e}")

    # OTC 上櫃：漲跌幅 + 成交量 + 收盤價（補充 TWSE 未包含的）
    missing = code_set - set(result)
    if missing:
        try:
            r = requests.get(
                "https://www.tpex.org.tw/openapi/v1/tpex_mainboard_quotes",
                headers={"Accept": "application/json"}, timeout=15, verify=False,
            )
            for item in (r.json() if r.ok else []):
                c = str(item.get("SecuritiesCompanyCode", "")).strip()
                if c not in missing:
                    continue
                result.setdefault(c, {})
                pct = _pct(item.get("Close"), item.get("Change"))
                if pct is not None:
                    result[c]["pct"] = pct
                result[c]["volume"] = _int(item.get("TradingShares", 0))
                p = _price(item.get("Close"))
                if p:
                    result[c]["price"] = p
        except Exception as e:
            print(f"  ⚠ OTC 漲跌幅/成交量查詢失敗：{e}")

    # TWSE 上市：已發行普通股數
    try:
        r = requests.get(
            "https://openapi.twse.com.tw/v1/opendata/t187ap03_L",
            headers={"Accept": "application/json"}, timeout=15, verify=False,
        )
        for item in (r.json() if r.ok else []):
            c = str(item.get("公司代號", "")).strip()
            if c not in code_set:
                continue
            result.setdefault(c, {})
            result[c]["issued"] = _int(item.get("已發行普通股數或TDR原股發行股數", 0))
    except Exception as e:
        print(f"  ⚠ TWSE 已發行股數查詢失敗：{e}")

    # OTC 上櫃：已發行普通股數（補充 TWSE 未包含的）
    missing_iss = {c for c in code_set if "issued" not in result.get(c, {})}
    if missing_iss:
        try:
            r = requests.get(
                "https://www.tpex.org.tw/openapi/v1/mopsfin_t187ap03_O",
                headers={"Accept": "application/json"}, timeout=15, verify=False,
            )
            for item in (r.json() if r.ok else []):
                c = str(item.get("SecuritiesCompanyCode", "")).strip()
                if c not in missing_iss:
                    continue
                result.setdefault(c, {})
                result[c]["issued"] = _int(item.get("IssueShares", 0))
        except Exception as e:
            print(f"  ⚠ OTC 已發行股數查詢失敗：{e}")

    n_pct = sum(1 for v in result.values() if "pct"    in v)
    n_vol = sum(1 for v in result.values() if "volume" in v)
    n_iss = sum(1 for v in result.values() if "issued" in v)
    print(f"  漲跌幅 {n_pct}/{len(codes)}  成交量 {n_vol}/{len(codes)}  已發行股數 {n_iss}/{len(codes)}")
    return result


# ── 主執行：抓取所有 ETF ──────────────────────────────────────────────

def _cleanup_old_snapshots(keep_days: int = 30) -> None:
    """刪除 etf_data/ 裡超過 keep_days 天的快照檔（以檔名日期判斷）"""
    cutoff = datetime.today() - timedelta(days=keep_days)
    cutoff_str = f"{cutoff.year - 1911}{cutoff.month:02d}{cutoff.day:02d}"
    removed = 0
    for p in DATA_DIR.glob("*.json"):
        # 檔名格式：00981A_1150724.json → 取底線後的數字部分
        stem = p.stem  # e.g. "00981A_1150724"
        parts = stem.rsplit("_", 1)
        if len(parts) != 2:
            continue
        date_part = parts[1]  # e.g. "1150724"
        if len(date_part) == 7 and date_part.isdigit() and date_part < cutoff_str:
            p.unlink()
            removed += 1
    if removed:
        print(f"  [ETF] 清理舊快照：刪除 {removed} 個超過 {keep_days} 天的檔案")


def run_all(force_fetch: bool = False) -> dict[str, dict]:
    """
    抓取所有設定 ETF，存快照，計算變動。
    force_fetch=True 可跳過快取強制重抓。
    回傳 {etf_code: {today, yesterday, changes}, "_price_changes": {...}}
    """
    _cleanup_old_snapshots(keep_days=30)
    results = {}
    for code, cfg in ETF_CONFIG.items():
        name   = cfg["name"]
        source = cfg["source"]
        param  = cfg["param"]
        print(f"  [ETF] {code} {name}...", end="", flush=True)

        # 今日快照已存在 → 直接讀快取，跳過 HTTP 請求
        # jpmorgan PCF 在盤中會更新，不快取以確保資料最新
        _skip_cache = source == "jpmorgan"
        if not force_fetch and not _skip_cache and _today_snapshot_exists(code):
            print(" ✓ 快取", end="")
            today_data = None  # 由 load_latest_snapshots 讀取
        else:
            if source == "fhtrust":
                today_data = fetch_holdings_fhtrust(code, param)
            elif source == "capitalfund":
                today_data = fetch_holdings_capitalfund(code, int(param))
            elif source == "cathayfund":
                today_data = fetch_holdings_cathayfund(code, str(param))
            elif source == "nomura":
                today_data = fetch_holdings_nomura(code, str(param))
            elif source == "allianzgi":
                today_data = fetch_holdings_allianzgi(code, str(param))
            elif source == "jpmorgan":
                today_data = fetch_holdings_jpmorgan(code, str(param))
            elif source == "fsitc":
                today_data = fetch_holdings_fsitc(code, str(param))
            elif source == "ctbcfund":
                today_data = fetch_holdings_ctbcfund(code, str(param))
            elif source == "megafund":
                today_data = fetch_holdings_megafund(code, int(param))
            elif source == "tsit":
                today_data = fetch_holdings_tsit(code, str(param))
            else:
                today_data = fetch_holdings(code, param)
            if not today_data or not today_data["holdings"]:
                print(" ❌ 無資料")
                continue
            save_snapshot(today_data)

        snapshots = load_latest_snapshots(code, 2)
        if not snapshots:
            print(" ❌ 快照讀取失敗")
            continue
        today     = snapshots[0]
        yesterday = snapshots[1] if len(snapshots) > 1 else None
        changes   = compute_changes(today, yesterday)

        added   = [c for c in changes if c["is_new"]]
        removed = [c for c in changes if c["is_removed"]]
        changed = [c for c in changes if not c["is_new"] and not c["is_removed"]
                   and c["delta_shares"] != 0]

        cached_mark = " [快取]" if (not force_fetch and today_data is None) else ""
        print(f"{cached_mark} {len(today['holdings'])} 檔 | "
              f"加碼 {len(added)} 減碼 {len(removed)} 調整 {len(changed)}")

        results[code] = {
            "today":     today,
            "yesterday": yesterday,
            "changes":   changes,
        }

    # 抓取所有持股當日漲跌幅（含已移除的股票，才能算變動估金額）
    if results:
        all_codes = list({
            h["code"]
            for res in results.values()
            for h in res["today"]["holdings"]
        } | {
            c["code"]
            for res in results.values()
            for c in res.get("changes", [])
            if c.get("is_removed")
        })
        results["_stock_data"] = fetch_stock_data(all_codes)

    return results


# ── HTML 產生 ─────────────────────────────────────────────────────────

def _fmt_shares(n: int) -> str:
    """11,840,000 → 11,840 張"""
    return f"{math.ceil(n / 1000):,}"


def _fmt_val(v: float) -> str:
    """267億 → 2,673 億"""
    yi = v / 1e8
    if yi >= 1:
        return f"{yi:.2f} 億"
    wan = v / 1e4
    return f"{wan:.0f} 萬"


def _delta_cell(delta: float | int, unit: str = "", order: float | int | None = None) -> str:
    if delta == 0 or delta is None:
        return "<td>-</td>"
    cls = "pos" if delta > 0 else "neg"
    sign = "+" if delta > 0 else ""
    v = order if order is not None else delta
    return f"<td class='{cls}' data-order='{v}'>{sign}{delta:,}{unit}</td>"


def _weight_cell(w: float) -> str:
    return f"<td>{w:.2f}%</td>" if w else "<td>-</td>"


def generate_etf_html(etf_results: dict[str, dict]) -> str:
    """產生主動ETF分頁的完整 HTML 內容（插入 <div id='etfTab'> 區塊）"""
    # 分離輔助資料與 ETF 結果
    stock_data: dict[str, dict] = etf_results.get("_stock_data", {}) or {}
    price_changes: dict[str, float | None] = {c: v.get("pct") for c, v in stock_data.items()}
    etf_results = {k: v for k, v in etf_results.items() if not k.startswith("_")}
    if not etf_results:
        return "<div class='no-data'>⚠️ ETF 資料尚未取得</div>"

    # ── 整合所有 ETF 的持股（股票總表）──────────────────────────────
    # 判斷是否有前日資料（任一 ETF 有 yesterday 就算有比較基準）
    has_prev = any(res["yesterday"] is not None for res in etf_results.values())

    # code → {name, funds, total_value, total_shares,
    #          net_delta_shares, buy_shares, sell_shares,
    #          net_delta_value, buy_value, sell_value, fund_count}
    stock_map: dict[str, dict] = {}
    for etf_code, res in etf_results.items():
        etf_name = res["today"]["etf_name"]
        nav = res["today"].get("nav_total") or 0
        etf_has_prev = res["yesterday"] is not None
        for row in res["changes"]:
            sc = row["code"]
            cur_shares = row["shares"]
            val = round(row["weight"] / 100 * nav) if (nav and cur_shares > 0) else 0
            # 估算每股價格：優先用 stock_data 真實收盤價，備援用持倉估算
            # 移除股（cur_shares=0）無法用今日持倉估價，改用前日 NAV 推估
            _sd_p = (stock_data.get(sc) or {}).get("price", 0)
            if _sd_p:
                price_est = _sd_p
            elif cur_shares > 0:
                price_est = val / cur_shares
            elif row.get("is_removed") and row.get("prev_shares", 0) > 0:
                _y_nav = (res.get("yesterday") or {}).get("nav_total") or 0
                _prev_val = round(row.get("prev_weight", 0) / 100 * _y_nav) if _y_nav else 0
                price_est = _prev_val / row["prev_shares"] if _prev_val else 0
            else:
                price_est = 0
            # 只有在有前日資料時才計算 delta（首日全部 is_new，不計算）
            if etf_has_prev and not row["is_new"] and not row["is_removed"]:
                delta_s = row["delta_shares"]
            elif etf_has_prev and (row["is_new"] or row["is_removed"]):
                delta_s = row["delta_shares"]  # 真正的新增/移除
            else:
                delta_s = 0  # 首日：全部視為無變動

            delta_val = round(delta_s * price_est) if (price_est and delta_s) else 0
            buy_s  = delta_s if delta_s > 0 else 0
            sell_s = -delta_s if delta_s < 0 else 0
            buy_v  = delta_val if delta_val > 0 else 0
            sell_v = -delta_val if delta_val < 0 else 0

            if sc not in stock_map:
                stock_map[sc] = {
                    "name": row["name"], "funds": [],
                    "total_value": 0, "total_shares": 0,
                    "net_delta_shares": 0, "buy_shares": 0, "sell_shares": 0,
                    "net_delta_value": 0, "buy_value": 0, "sell_value": 0,
                    "fund_count": 0, "per_etf": {},
                }
            stock_map[sc]["funds"].append(etf_name)
            stock_map[sc]["total_value"]      += val
            stock_map[sc]["total_shares"]     += cur_shares
            stock_map[sc]["net_delta_shares"] += delta_s
            stock_map[sc]["buy_shares"]       += buy_s
            stock_map[sc]["sell_shares"]      += sell_s
            stock_map[sc]["net_delta_value"]  += delta_val
            stock_map[sc]["buy_value"]        += buy_v
            stock_map[sc]["sell_value"]       += sell_v
            if cur_shares > 0:
                stock_map[sc]["fund_count"]   += 1
            stock_map[sc]["per_etf"][etf_code] = {
                "en": etf_name, "s": cur_shares, "d": delta_s, "dv": delta_val, "w": row["weight"],
            }

    # 依總持有估值排序；持有估值=0（僅1股掛號）排最後
    sorted_stocks = sorted(stock_map.items(),
                           key=lambda x: x[1]["total_value"], reverse=True)

    # ── 股票總表 ─────────────────────────────────────────────────────
    _SORT_NONE = -999999999  # "-" 欄位排序權重（置底）

    def _dash(): return f"<td class='text-muted' data-order='{_SORT_NONE}'>-</td>"

    stock_rows = ""
    summary_stocks = []
    for sc, sd in sorted_stocks:
        if sd["total_shares"] == 0:
            continue
        is_nominal = sd["total_value"] < 10_000  # 持有 1 股掛號位置

        has_delta   = has_prev and sd["net_delta_shares"] != 0
        net_val     = sd["net_delta_value"]
        net_lots    = math.ceil(sd["net_delta_shares"] / 1000) if sd["net_delta_shares"] >= 0 else -math.ceil(abs(sd["net_delta_shares"]) / 1000)
        if has_delta:
            _nv_cls  = "pos" if net_val > 0 else "neg"
            _nv_sign = "+" if net_val > 0 else "-"
            net_amt_cell = f"<td class='{_nv_cls}' data-order='{net_val}'>{_nv_sign}{abs(net_val)/1e8:.2f} 億</td>"
        elif has_prev:
            net_amt_cell = f"<td class='text-muted' data-order='0'>0.00 億</td>"
        else:
            net_amt_cell = _dash()
        if has_delta:
            net_lot_cell = _delta_cell(net_lots, "", order=net_lots)
        elif has_prev:
            net_lot_cell = f"<td class='text-muted' data-order='0'>0</td>"
        else:
            net_lot_cell = _dash()

        buy_v  = sd["buy_value"]
        sell_v = sd["sell_value"]
        buy_cell  = (f"<td class='pos' data-order='{buy_v}'>{buy_v/1e8:.2f} 億</td>"
                     if (has_prev and sd["buy_shares"]) else _dash())
        sell_cell = (f"<td class='neg' data-order='{sell_v}'>-{sell_v/1e8:.2f} 億</td>"
                     if (has_prev and sd["sell_shares"]) else _dash())

        fund_count  = sd["fund_count"]
        tot_val     = sd["total_value"]
        tot_shares  = sd["total_shares"]
        nominal_style = "opacity:.5;font-size:.85em;" if is_nominal else ""
        val_display = '<span class="text-muted">-</span>' if is_nominal else f"{tot_val/1e8:.2f} 億"

        # 漲跌%
        chg_pct = price_changes.get(sc)
        if chg_pct is not None:
            cls = "pos" if chg_pct > 0 else ("neg" if chg_pct < 0 else "")
            sign = "+" if chg_pct > 0 else ""
            chg_cell = f"<td class='{cls}' data-order='{chg_pct}'>{sign}{chg_pct:.2f}%</td>"
        else:
            chg_cell = _dash()

        # 股本比（淨變動張數／已發行股數）
        sdi = stock_data.get(sc) or {}
        issued  = sdi.get("issued", 0)
        volume  = sdi.get("volume", 0)
        net_raw = sd["net_delta_shares"]  # in shares

        _sbc_save = None
        if has_delta and issued:
            _sbc_r = net_raw / issued * 100
            _sbc_save = round(_sbc_r, 3)
            _sbc_s = "+" if _sbc_r > 0 else ""
            _sbc_c = "pos" if _sbc_r > 0 else ("neg" if _sbc_r < 0 else "")
            sbc_stock_cell = f"<td class='{_sbc_c}' data-order='{_sbc_r}'>{_sbc_s}{abs(_sbc_r):.3f}%</td>"
        elif has_prev and issued:
            sbc_stock_cell = f"<td class='text-muted' data-order='0'>0.000%</td>"
        else:
            sbc_stock_cell = _dash()

        # 量比（淨變動張數／當日成交量）
        _vr_save = None
        if has_delta and volume:
            _vr_r = net_raw / volume * 100
            _vr_save = round(_vr_r, 2)
            _vr_s = "+" if _vr_r > 0 else ""
            _vr_c = "pos" if _vr_r > 0 else ("neg" if _vr_r < 0 else "")
            vr_cell = f"<td class='{_vr_c}' data-order='{_vr_r}'>{_vr_s}{abs(_vr_r):.2f}%</td>"
        elif has_prev and volume:
            vr_cell = f"<td class='text-muted' data-order='0'>0.00%</td>"
        else:
            vr_cell = _dash()

        summary_stocks.append({
            "c": sc, "n": sd["name"],
            "ndv": sd["net_delta_value"], "ndl": net_lots,
            "sbc": _sbc_save, "chg": price_changes.get(sc), "vr": _vr_save,
            "ts": tot_shares, "tv": tot_val, "fc": fund_count,
            "bv": sd["buy_value"], "sv": sd["sell_value"],
            "etfs": [
                {"ec": ec, "en": v["en"], "s": v["s"], "w": v["w"],
                 "dl": (math.ceil(v["d"]/1000) if v["d"] >= 0 else -math.ceil(abs(v["d"])/1000)),
                 "dv": v["dv"]}
                for ec, v in sd.get("per_etf", {}).items()
            ],
        })

        sname_safe = sd['name'].replace("'", "&#39;")
        stock_rows += (
            f"<tr style='{nominal_style}cursor:pointer' onclick=\"filterEtfStock('{sc}', '{sname_safe}')\">"
            f"<td data-order='{sc}'><b style='color:#4fc3f7'>{sc}</b> {sd['name']}</td>"
            f"{net_amt_cell}"
            f"{net_lot_cell}"
            f"{sbc_stock_cell}"
            f"{chg_cell}"
            f"{vr_cell}"
            f"<td data-order='{tot_shares}'>{_fmt_shares(tot_shares)}</td>"
            f"<td data-order='{tot_val}'>{val_display}</td>"
            f"<td class='text-center' data-order='{fund_count}'>{fund_count}</td>"
            f"{buy_cell}"
            f"{sell_cell}"
            f"</tr>"
        )

    active_count = len([s for s, d in sorted_stocks if d["total_shares"] > 0])
    total_buy_all  = sum(d["buy_value"]  for _, d in sorted_stocks if has_prev)
    total_sell_all = sum(d["sell_value"] for _, d in sorted_stocks if has_prev)
    net_all        = total_buy_all - total_sell_all


    net_all_cls    = "color:#ff6b6b" if net_all < 0 else "color:#4ecdc4"
    net_all_sign   = "+" if net_all > 0 else ("-" if net_all < 0 else "")
    summary_bar = f"""
<div style="background:#1a1a2e;color:#fff;display:flex;gap:0;border-bottom:1px solid #333" class="px-0">
  <div style="flex:1;padding:10px 20px;border-right:1px solid #333">
    <span style="color:#aaa;font-size:.8rem">買進</span>
    <span style="color:#4ecdc4;font-size:1.4rem;font-weight:700;margin-left:12px">{total_buy_all/1e8:.0f} 億</span>
  </div>
  <div style="flex:1;padding:10px 20px;border-right:1px solid #333">
    <span style="color:#aaa;font-size:.8rem">賣出</span>
    <span style="color:#ff6b6b;font-size:1.4rem;font-weight:700;margin-left:12px">{total_sell_all/1e8:.0f} 億</span>
  </div>
  <div style="flex:1;padding:10px 20px">
    <span style="color:#aaa;font-size:.8rem">買賣超</span>
    <span style="{net_all_cls};font-size:1.4rem;font-weight:700;margin-left:12px">{net_all_sign}{abs(net_all)/1e8:.0f} 億</span>
  </div>
</div>""" if has_prev else ""
    stock_table = f"""
<div class="card mb-3">
  <div class="card-header px-3 py-2 fw-bold d-flex align-items-center flex-wrap gap-2">
    <span id="etfStockTitle">股票總表</span>
    <small class="text-muted fw-normal">點股票，看哪些 ETF 持有</small>
    <span id="etfStockCount" class="fw-normal text-muted" style="font-size:.85rem">共 {active_count} 檔</span>
    <span id="etfStockBreadcrumb" class="ms-auto text-muted d-none" style="font-size:.8rem">
      股票總表 → 持股明細 → 各基金買賣概況
    </span>
  </div>
  {summary_bar}
  <div id="etfStockFilterBar" class="px-3 pt-2 pb-1 border-bottom d-none">
    <div class="d-flex align-items-center gap-2">
      <input id="etfStockSearchInput" class="form-control form-control-sm" style="max-width:160px" placeholder="股票代號" oninput="onEtfStockInput(this.value)">
      <button class="btn btn-sm btn-outline-secondary py-0" onclick="clearEtfStockFilter()">清除篩選</button>
    </div>
  </div>
  <div class="table-responsive">
  <table class="table table-hover table-sm mb-0" id="etfStockTable">
    <thead><tr>
      <th>標的</th>
      <th title="淨變動估金額" style="cursor:help">淨額</th>
      <th title="淨變動張數" style="cursor:help">張數</th>
      <th title="淨變動張數／已發行股數" style="cursor:help">股本比</th>
      <th title="當日漲跌幅" style="cursor:help">漲跌%</th>
      <th title="淨變動張數／當日成交量" style="cursor:help">量比</th>
      <th>持有張</th><th>持有估值</th>
      <th title="持有本股的基金數量" style="cursor:help">基金</th>
      <th title="買進估金額" style="cursor:help">買進</th>
      <th title="賣出估金額" style="cursor:help">賣出</th>
    </tr></thead>
    <tbody>{stock_rows}</tbody>
  </table>
  </div>
</div>"""

    # ── 本日持股變動明細 ──────────────────────────────────────────────
    _S0 = -999999999  # "-" 欄位排序置底值

    total_holdings = sum(
        len([c for c in res["changes"] if c["shares"] > 0])
        for res in etf_results.values()
    )

    # 蒐集所有 ETF 的所有持股（有前日資料才算 delta，首日顯示 "-"）
    all_list = []
    for etf_code, res in etf_results.items():
        has_prev = res["yesterday"] is not None
        etf_name = res["today"]["etf_name"]
        nav = res["today"].get("nav_total") or 0
        for row in res["changes"]:
            cur_shares = row["shares"]
            val = round(row["weight"] / 100 * nav) if (nav and cur_shares > 0) else 0
            if has_prev:
                _sd_price = (stock_data.get(row["code"]) or {}).get("price", 0)
                if _sd_price:
                    price_est = _sd_price
                elif cur_shares > 0:
                    price_est = val / cur_shares
                elif row.get("is_removed") and row.get("prev_shares", 0) > 0:
                    _y_nav = (res.get("yesterday") or {}).get("nav_total") or 0
                    _prev_val = round(row.get("prev_weight", 0) / 100 * _y_nav) if _y_nav else 0
                    price_est = _prev_val / row["prev_shares"] if _prev_val else 0
                else:
                    price_est = 0
                delta_val = round(row["delta_shares"] * price_est) if price_est else 0
                is_changed = row["delta_shares"] != 0
            else:
                delta_val = None   # 首日：無可比對
                is_changed = False
            all_list.append({
                "etf_code": etf_code,
                "etf_name": etf_name,
                "row": row,
                "val": val,
                "delta_val": delta_val,
                "is_changed": is_changed,
                "has_prev": has_prev,
            })

    # 預設排序：有變動的先（依 |delta_val| 降冪），無變動的排後（依 val 降冪）
    all_list.sort(key=lambda x: (0 if x["is_changed"] else 1, -abs(x["delta_val"] or 0), -x["val"]))
    change_count = sum(1 for x in all_list if x["is_changed"])

    # ── 為歷史 summary 收集 ETF 級別資料 ────────────────────────────────
    summary_etfs = []
    for _ec, _res in etf_results.items():
        _today_d   = _res["today"]
        _hp        = _res["yesterday"] is not None
        _nav       = _today_d.get("nav_total") or 0
        _etf_holdings = []
        _nbv = _sbv = 0.0
        for _row in _res["changes"]:
            _sc   = _row["code"]
            _cs   = _row["shares"]
            _val  = round(_row["weight"] / 100 * _nav) if (_nav and _cs > 0) else 0
            _p    = (stock_data.get(_sc) or {}).get("price", 0) or (_val / _cs if _cs > 0 else 0)
            _ds   = _row["delta_shares"] if _hp else 0
            _dl   = (math.ceil(_ds/1000) if _ds >= 0 else -math.ceil(abs(_ds)/1000))
            _dv   = round(_ds * _p) if (_p and _hp) else 0
            _sbc  = round(_ds / (stock_data.get(_sc) or {}).get("issued", 0) * 100, 3) \
                    if (_hp and _ds and (stock_data.get(_sc) or {}).get("issued")) else None
            _chg  = price_changes.get(_sc)
            if _dv > 0: _nbv += abs(_dv)
            elif _dv < 0: _sbv += abs(_dv)
            _etf_holdings.append({
                "c": _sc, "n": _row["name"],
                "s": _cs, "w": _row["weight"], "v": _val,
                "dl": _dl, "dv": _dv, "sbc": _sbc, "chg": _chg,
                "is_new": _row["is_new"], "is_rm": _row["is_removed"],
                "changed": _hp and (_row["delta_shares"] != 0),
                "hp": _hp,
            })
        summary_etfs.append({
            "ec": _ec, "en": _today_d["etf_name"],
            "nav": _nav, "date": _today_d.get("date", ""),
            "nbv": _nbv, "sbv": _sbv, "hp": _hp,
            "holdings": _etf_holdings,
        })

    # ── 為歷史 summary 收集 change_list ──────────────────────────────────
    summary_changes = []
    for _item in all_list:
        _row  = _item["row"]
        _sc   = _row["code"]
        _ds   = _row["delta_shares"]
        _dl   = math.ceil(_ds/1000) if _ds >= 0 else -math.ceil(abs(_ds)/1000)
        _iss  = (stock_data.get(_sc) or {}).get("issued", 0)
        _sbc  = round(_ds / _iss * 100, 3) if (_iss and _item["has_prev"] and _ds) else None
        summary_changes.append({
            "ec": _item["etf_code"], "en": _item["etf_name"],
            "c": _sc, "n": _row["name"],
            "w": _row["weight"], "s": _row["shares"], "v": _item["val"],
            "dv": _item["delta_val"], "dl": _dl, "sbc": _sbc,
            "chg": price_changes.get(_sc),
            "is_new": _row["is_new"], "is_rm": _row["is_removed"],
            "changed": _item["is_changed"], "hp": _item["has_prev"],
        })

    # ── 存今日彙整快照（供歷史日期下拉使用）──────────────────────────
    # 用資料本身的日期（非日曆今天），非交易日執行時會蓋回同一個交易日檔，不新增空頁
    _data_dates = [res["today"]["date"] for res in etf_results.values()
                   if res.get("today") and res["today"].get("date")]
    if _data_dates:
        _ref_date = max(_data_dates)          # e.g. "115/08/21"
        _today_suffix = _ref_date.replace("/", "")  # → "1150821"
    else:
        _today_suffix = f"{datetime.today().year - 1911}{datetime.today().month:02d}{datetime.today().day:02d}"
    _summary_path = DATA_DIR / f"summary_{_today_suffix}.json"
    try:
        _summary_path.write_text(json.dumps({
            "date_str": _today_suffix,
            "has_prev": has_prev,
            "total_buy": total_buy_all,
            "total_sell": total_sell_all,
            "stocks": summary_stocks,
            "etfs": summary_etfs,
            "changes": summary_changes,
        }, ensure_ascii=False), encoding="utf-8")
    except Exception as _e:
        print(f"  ⚠ ETF summary 存檔失敗：{_e}")

    # 掃描可用歷史日期（存檔後再掃，才包含今天）
    _avail_dates = []
    for _p in sorted(DATA_DIR.glob("summary_*.json"), reverse=True):
        _dp = _p.stem[len("summary_"):]
        if len(_dp) == 7 and _dp.isdigit():
            _avail_dates.append({"key": _dp, "label": f"{_dp[:3]}/{_dp[3:5]}/{_dp[5:7]}"})
    _dates_json = json.dumps(_avail_dates, ensure_ascii=False)

    change_rows = ""
    for item in all_list:
        etf_code  = item["etf_code"]
        etf_name  = item["etf_name"]
        row       = item["row"]
        val       = item["val"]
        delta_val = item["delta_val"]
        is_chg    = item["is_changed"]
        has_prev  = item["has_prev"]
        ds_raw = row["delta_shares"]
        delta_s = math.ceil(ds_raw / 1000) if ds_raw >= 0 else -math.ceil(abs(ds_raw) / 1000)
        chg_pct   = price_changes.get(row["code"])

        tag = ""
        if has_prev and row["is_new"]:
            tag = "<span class='badge bg-success ms-1'>新增</span>"
        elif has_prev and row["is_removed"]:
            tag = "<span class='badge bg-danger ms-1'>移除</span>"

        # 變動估金額（首日全 "-"，單位統一億）
        if not has_prev or delta_val is None:
            dv_cell = f"<td class='text-muted' data-order='{_S0}'>-</td>"
        elif delta_val:
            cls  = "pos" if delta_val > 0 else "neg"
            sign = "+" if delta_val > 0 else "-"
            dv_cell = f"<td class='{cls}' data-order='{delta_val}'>{sign}{abs(delta_val)/1e8:.2f} 億</td>"
        else:
            dv_cell = f"<td class='text-muted' data-order='0'>0.00 億</td>"

        # △張（首日全 "-"）
        if not has_prev:
            ds_cell = f"<td class='text-muted' data-order='{_S0}'>-</td>"
        elif delta_s:
            ds_cell = _delta_cell(delta_s, "", order=delta_s)
        else:
            ds_cell = f"<td class='text-muted' data-order='0'>0</td>"

        # △股本比（變動張數／已發行股數）
        _iss = (stock_data.get(row["code"]) or {}).get("issued", 0)
        if has_prev and ds_raw and _iss:
            _cr = ds_raw / _iss * 100
            _cs = "+" if _cr > 0 else ""
            _cc = "pos" if _cr > 0 else ("neg" if _cr < 0 else "")
            sbc_cell = f"<td class='{_cc}' data-order='{_cr}'>{_cs}{abs(_cr):.3f}%</td>"
        else:
            sbc_cell = f"<td class='text-muted' data-order='{_S0}'>-</td>"

        # 日漲跌
        if chg_pct is not None:
            cls = "pos" if chg_pct > 0 else ("neg" if chg_pct < 0 else "")
            sign = "+" if chg_pct > 0 else ""
            chg_cell = f"<td class='{cls}' data-order='{chg_pct}'>{sign}{chg_pct:.2f}%</td>"
        else:
            chg_cell = f"<td class='text-muted' data-order='{_S0}'>-</td>"

        chg_flag = "1" if is_chg else "0"
        change_rows += (
            f"<tr data-changed='{chg_flag}' data-stock='{row['code']}'>"
            f"<td data-search='{etf_code}'>"
            f"<small class='text-muted'>{etf_code}</small> {etf_name}</td>"
            f"<td><b style='color:#4fc3f7'>{row['code']}</b> {row['name']}{tag}</td>"
            f"<td data-order='{row['weight']}'>{row['weight']:.2f}%</td>"
            f"<td data-order='{row['shares']}'>{_fmt_shares(row['shares'])}</td>"
            f"<td data-order='{val}'>{_fmt_val(val) if val else '-'}</td>"
            f"{dv_cell}{ds_cell}{sbc_cell}{chg_cell}"
            f"</tr>"
        )

    change_table = f"""
<div class="card mb-3">
  <div class="card-header px-3 py-2 fw-bold d-flex align-items-center flex-wrap gap-2">
    <span id="etfChangeTitle">本日變動持股明細</span>
    <span id="etfChangeCount" class="fw-normal text-muted" style="font-size:.85rem">
      變動 {change_count} 筆 / 全部持股 {total_holdings} 筆
    </span>
    <span id="etfFundChip" class="ms-1"></span>
    <span id="etfStockChip" class="ms-1"></span>
    <div class="ms-auto d-flex gap-2">
      <button id="etfClearFundBtn"  class="btn btn-sm btn-outline-secondary py-0 d-none" onclick="clearEtfFund()">清除基金</button>
      <button id="etfClearStockBtn" class="btn btn-sm btn-outline-secondary py-0 d-none" onclick="clearEtfStock()">清除股票</button>
    </div>
  </div>
  <div class="px-3 pt-2 pb-1 border-bottom d-flex align-items-center gap-3">
    <button id="etfChangedOnlyBtn" class="btn btn-sm btn-primary py-0 px-2"
            onclick="toggleEtfChangedOnly()">本日有加減碼</button>
    <small id="etfSortHint" class="text-muted">
      依變動估金額降冪排序；共 {change_count} 筆變動
    </small>
  </div>
  <div class="table-responsive">
  <table class="table table-hover table-sm mb-0" id="etfChangeTable">
    <thead><tr>
      <th>基金</th><th>股票</th><th>比例</th><th>張數</th>
      <th>估市值</th>
      <th title="持股變動對應的估算金額" style="cursor:help">變動估金額</th>
      <th title="變動張數" style="cursor:help">△張</th>
      <th title="變動張數／已發行股數" style="cursor:help">△股本比</th>
      <th title="當日漲跌幅" style="cursor:help">日漲跌</th>
    </tr></thead>
    <tbody>{change_rows}</tbody>
  </table>
  </div>
</div>"""

    # ── 各基金買賣概況 ────────────────────────────────────────────────
    fund_rows = ""
    for etf_code, res in etf_results.items():
        today = res["today"]
        nav = today.get("nav_total") or 0
        nav_unit = today.get("nav_per_unit") or 0
        date_str = today.get("date", "")

        added   = [c for c in res["changes"] if c["is_new"]]
        removed = [c for c in res["changes"] if c["is_removed"]]
        up      = [c for c in res["changes"] if not c["is_new"] and not c["is_removed"] and c["delta_shares"] > 0]
        down    = [c for c in res["changes"] if not c["is_new"] and not c["is_removed"] and c["delta_shares"] < 0]

        # 買方：新增 + 加碼（依 delta 降冪），賣方：減碼 + 移除（依 |delta| 降冪）
        all_buys  = sorted(added + up,      key=lambda c: c["delta_shares"],       reverse=True)
        all_sells = sorted(removed + down,  key=lambda c: abs(c["delta_shares"]),  reverse=True)

        def _stock_price(c):
            p = (stock_data.get(c["code"]) or {}).get("price", 0)
            if p:
                return p
            # 備援：用持倉比例反推估價（移除股票此值為 0）
            sh = c["shares"]
            if sh and nav:
                return round(c["weight"] / 100 * nav / sh)
            return 0

        def _fmt_item_buy(c):
            lots = c["delta_shares"] // 1000
            p    = _stock_price(c)
            val  = c["delta_shares"] * p if p else 0
            v_str = f" / +{val/1e8:.2f}億" if val else ""
            return f"<span style='color:#4fc3f7'>{c['code']}</span>{c['name']} +{lots}張{v_str}"

        def _fmt_item_sell(c):
            lots = abs(c["delta_shares"]) // 1000
            p    = _stock_price(c)
            val  = abs(c["delta_shares"]) * p if p else 0
            v_str = f" / -{val/1e8:.2f}億" if val else ""
            return f"<span style='color:#4fc3f7'>{c['code']}</span><span style='color:#fff'>{c['name']}</span> -{lots}張{v_str}"

        def _truncate_items(items_fn, lst):
            """≤4 檔全列，>4 列前3檔+等X檔；每行放2檔。"""
            if not lst:
                return []
            n = len(lst)
            if n <= 4:
                items = [items_fn(c) for c in lst]
            else:
                items = [items_fn(c) for c in lst[:3]] + [f"等{n}檔"]
            # 每行放2檔
            sep = '<span style="color:#4a5568;margin:0 4px">·</span>'
            lines = []
            for i in range(0, len(items), 2):
                pair = items[i:i+2]
                lines.append(sep.join(pair))
            return lines
        buy_items  = _truncate_items(_fmt_item_buy,  all_buys)
        sell_items = _truncate_items(_fmt_item_sell, all_sells)
        buy_str  = "<br>".join(buy_items)  if buy_items  else "-"
        sell_str = "<br>".join(sell_items) if sell_items else "-"

        def _dv(c):
            p = _stock_price(c)
            return abs(c["delta_shares"]) * p if p else 0

        total_buy_val  = sum(_dv(c) for c in all_buys)
        total_sell_val = sum(_dv(c) for c in all_sells)
        total_add = sum(c["delta_shares"] // 1000     for c in all_buys)
        total_rm  = sum(abs(c["delta_shares"]) // 1000 for c in all_sells)
        holdings_n  = len([c for c in res["changes"] if c["shares"] > 0])
        changed_n   = len(all_buys) + len(all_sells)

        cfg = ETF_CONFIG.get(etf_code, {})
        if cfg.get("source") == "fhtrust":
            fund_url = f"https://www.fhtrust.com.tw/ETF/etf_detail/{cfg['param']}#stockhold"
        elif cfg.get("source") == "capitalfund":
            fund_url = f"https://www.capitalfund.com.tw/etf/product/detail/{cfg['param']}/portfolio"
        elif cfg.get("source") == "cathayfund":
            fund_url = f"https://www.cathaysite.com.tw/ETF/detail/E{cfg['param']}?tab=etf3"
        elif cfg.get("source") == "nomura":
            fund_url = f"https://www.nomurafunds.com.tw/ETFWEB/product-description?fundNo={cfg['param']}&tab=Shareholding"
        elif cfg.get("source") == "allianzgi":
            fund_url = f"https://etf.allianzgi.com.tw/etf-info/{cfg['param']}?tab=4"
        elif cfg.get("source") == "jpmorgan":
            fund_url = f"https://am.jpmorgan.com/tw/zh/asset-management/twetf/products/jpmorgan-taiwan-taiwan-equity-high-income-active-etf-{cfg['param'].lower()}#/portfolio"
        elif cfg.get("source") == "fsitc":
            fund_url = f"https://www.fsitc.com.tw/FundDetail.aspx?ID={cfg['param']}#TabLinkdivEditTab3"
        elif cfg.get("source") == "ctbcfund":
            fund_url = f"https://www.ctbcinvestments.com/Etf/{etf_code}/Combination"
        elif cfg.get("source") == "megafund":
            fund_url = f"https://www.megafunds.com.tw/MEGA/etf/etf_product.aspx?id={cfg['param']}"
        elif cfg.get("source") == "tsit":
            fund_url = f"https://www.tsit.com.tw/ETF/Home/ETFSeriesDetail/{cfg['param']}"
        else:
            fund_url = f"https://www.ezmoney.com.tw/ETF/Fund/Info/{etf_code}?fundCode={cfg.get('param','')}"
        etf_name_safe = today['etf_name'].replace("'", "&#39;")
        prev_date = res["yesterday"]["date"] if res["yesterday"] else None
        net_val  = total_buy_val - total_sell_val
        net_sign = "+" if net_val > 0 else ("-" if net_val < 0 else "")
        net_cls  = "neg" if net_val < 0 else ""
        net_main = f"{net_sign}{abs(net_val)/1e8:.2f} 億" if net_val else "0"
        buy_sub  = f"+{total_buy_val/1e8:.2f} 億"  if total_buy_val  else "-"
        sell_sub = f"-{total_sell_val/1e8:.2f} 億" if total_sell_val else "-"
        fund_rows += f"""
<tr style="cursor:pointer" onclick="filterEtfFund('{etf_code}', '{etf_name_safe}')">
  <td data-order='{etf_code}'>
    <a href='{fund_url}' target='_blank' onclick="event.stopPropagation()"><b>{etf_code}</b> {today['etf_name']}</a>
  </td>
  <td data-order='{net_val}'><span class='{net_cls}'>{net_main}</span><br>
    <small class='fw-normal' style='font-size:.75rem;color:inherit'>買{buy_sub} / 賣{sell_sub}</small>
  </td>
  <td data-order='{total_add}' style='white-space:normal'>{buy_str}</td>
  <td class='{("neg" if all_sells else "")}' data-order='{total_rm}' style='white-space:normal'>{sell_str}</td>
  <td data-order='{changed_n}'>
    <span style='font-size:1rem;font-weight:700'>{changed_n}</span> <small class='text-muted'>變動</small><br>
    <small class='text-muted'>{holdings_n} 持股</small>
  </td>
  <td data-order='{nav}'>{_fmt_val(nav) if nav else "-"}</td>
  <td data-order='{date_str}'>{date_str}{"<br><small class='text-muted'>（前次：" + prev_date + "）</small>" if prev_date else ""}</td>
  <td><a href='{fund_url}' target='_blank' onclick="event.stopPropagation()" class='btn btn-outline-secondary btn-sm py-0'>官方頁</a></td>
</tr>"""

    fund_table = f"""
<div class="card mb-3">
  <div class="card-header px-3 py-2 fw-bold">各基金買賣概況　{len(etf_results)} 檔</div>
  <div class="table-responsive">
  <table class="table table-hover table-sm mb-0" id="etfFundTable">
    <thead><tr>
      <th>基金</th><th>淨買賣</th><th>加碼</th><th>減碼</th><th>變動</th><th>規模</th><th>日期</th><th></th>
    </tr></thead>
    <tbody>{fund_rows}</tbody>
  </table>
  </div>
</div>"""

    # 找出最新基準日（最大日期），與其不同的 ETF 視為未更新
    all_dates = [res["today"]["date"] for res in etf_results.values() if res["today"].get("date")]
    ref_date = max(all_dates) if all_dates else ""

    # 更新時間：最新 ETF 資料日期（民國→西元）+ 最晚 fetch_time
    all_fetch_times = [
        res["today"].get("fetch_time", "")
        for res in etf_results.values()
        if res["today"].get("date") == ref_date and res["today"].get("fetch_time")
    ]
    latest_hm = max(all_fetch_times) if all_fetch_times else ""
    if ref_date:
        _parts = ref_date.split("/")
        try:
            _date_str = f"{int(_parts[0]) + 1911}/{_parts[1]}/{_parts[2]}"
        except Exception:
            _date_str = ref_date
        update_time = f"{_date_str} {latest_hm}".strip()
    else:
        update_time = datetime.now().strftime("%Y/%m/%d %H:%M")
    stale = [(code, res["today"]["etf_name"], res["today"]["date"])
             for code, res in etf_results.items()
             if res["today"].get("date", "") < ref_date]
    if stale:
        stale_items = "".join(
            f"<span style='margin-right:8px'><b>{code}</b> <span class='text-muted'>{date}</span></span>"
            for code, name, date in stale
        )
        stale_badge = f"""<div class='ms-auto d-flex align-items-center gap-1 flex-wrap' style='font-size:.78rem'>
          <span style='color:#cf222e;font-weight:600'>⚠ 未更新</span>
          {stale_items}
        </div>"""
    else:
        stale_badge = ""

    _dropdown_html = ""
    if _avail_dates:
        _opts = "".join(
            '<option value="' + d["key"] + '">' + d["label"] + '</option>'
            for d in _avail_dates
        )
        _dropdown_html = (
            '<select id="etfDateSelect" class="form-select form-select-sm d-inline-block" '
            'style="width:auto;min-width:120px;background-color:#2a2a3e;color:#e0e0e0;border-color:#555" '
            'onchange="etfSelectDate(this.value)">'
            + _opts + '</select>'
        )

    return f"""
<div class="container-fluid px-3 py-2">
  <div class="d-flex align-items-center mb-2 gap-2 flex-wrap">
    <h5 class="mb-0">主動ETF持股追蹤</h5>
    <small class="text-muted">更新：{update_time}</small>
    {_dropdown_html}
    {stale_badge}
  </div>
  <div id="etfCurrentView">
    {stock_table}
    {change_table}
    {fund_table}
  </div>
  <div id="etfHistoryView" style="display:none"></div>
</div>
<script>
(function(){{
  var AVAIL={_dates_json};
  var LIVE_KEY=AVAIL.length?AVAIL[0].key:null;
  var S0=-999999999;
  var _hData=null, _hDTs={{}};
  var _hFundFilter=null, _hStockFilter=null, _hChangedOnly=true;

  // ── format helpers ──
  function fA(v,hp){{
    if(v===null||v===undefined||v===0) return hp?'0.00 億':'-';
    var a=Math.abs(v),s=v>0?'+':'-';
    return a>=1e8?s+(a/1e8).toFixed(2)+' 億':s+(a/1e4).toFixed(0)+' 萬';
  }}
  function fL(v){{return v?(v>0?'+':'')+v.toLocaleString():null;}}
  function fS(v){{return Math.ceil(v/1000).toLocaleString()+' 張';}}
  function fV(v){{return(!v||v<10000)?'-':(v/1e8).toFixed(2)+' 億';}}
  function cl(v){{return v>0?'pos':v<0?'neg':'';}}
  function fP(v,d){{return v!=null?(v>0?'+':'')+Math.abs(v).toFixed(d||2)+'%':null;}}

  // ── destroy history DataTables ──
  function _destroyHDTs(){{
    Object.keys(_hDTs).forEach(function(id){{
      if($.fn.DataTable.isDataTable('#'+id)){{
        $('#'+id).DataTable().destroy();
      }}
      delete _hDTs[id];
    }});
  }}

  // ── init history DataTable ──
  function _initDT(id, orderCol, colDefs){{
    if(!$.fn.DataTable||!document.getElementById(id)) return;
    var dt=$('#'+id).DataTable({{
      paging:false, autoWidth:false,
      order:[[orderCol,'desc']],
      language:{{search:'搜尋：',info:'共 _TOTAL_ 筆',zeroRecords:'無資料'}},
      columnDefs: colDefs||[],
    }});
    _hDTs[id]=dt;
    return dt;
  }}

  // ── build summary bar ──
  function _bar(tb,ts_,hp){{
    if(!hp) return '';
    var net=(tb||0)-(ts_||0),nc=net<0?'#ff6b6b':'#4ecdc4',ns=net>0?'+':net<0?'-':'';
    return '<div style="background:#1a1a2e;color:#fff;display:flex;gap:0;border-bottom:1px solid #333" class="px-0">'+
      '<div style="flex:1;padding:10px 20px;border-right:1px solid #333"><span style="color:#aaa;font-size:.8rem">買進</span>'+
      '<span style="color:#4ecdc4;font-size:1.4rem;font-weight:700;margin-left:12px">'+((tb||0)/1e8).toFixed(0)+' 億</span></div>'+
      '<div style="flex:1;padding:10px 20px;border-right:1px solid #333"><span style="color:#aaa;font-size:.8rem">賣出</span>'+
      '<span style="color:#ff6b6b;font-size:1.4rem;font-weight:700;margin-left:12px">'+((ts_||0)/1e8).toFixed(0)+' 億</span></div>'+
      '<div style="flex:1;padding:10px 20px"><span style="color:#aaa;font-size:.8rem">買賣超</span>'+
      '<span style="color:'+nc+';font-size:1.4rem;font-weight:700;margin-left:12px">'+ns+(Math.abs(net)/1e8).toFixed(0)+' 億</span></div></div>';
  }}

  // ── build stock table HTML ──
  function _buildStockRows(stocks,hp){{
    return (stocks||[]).filter(function(s){{return s.ts>0;}}).map(function(s){{
      var ndv=s.ndv||0, nl=s.ndl||0;
      var ndvC=ndv?cl(ndv):'text-muted', ndvF=ndv?fA(ndv,hp):(hp?'0.00 億':'-'), ndvO=ndv||(hp?0:S0);
      var nlF=fL(nl), nlC=nlF?cl(nl):'text-muted', nlO=nl||(hp?0:S0);
      var sbcF=fP(s.sbc,3)||(hp?'0.000%':'-'), sbcC=s.sbc!=null?cl(s.sbc):'text-muted', sbcO=s.sbc!=null?s.sbc:(hp?0:S0);
      var chgF=fP(s.chg,2)||'-', chgC=s.chg!=null?cl(s.chg):'text-muted', chgO=s.chg!=null?s.chg:S0;
      var vrF=fP(s.vr,2)||(hp?'0.00%':'-'), vrC=s.vr!=null?cl(s.vr):'text-muted', vrO=s.vr!=null?s.vr:(hp?0:S0);
      var bvO=s.bv||S0, svO=s.sv||S0;
      var nn=s.n.replace(/"/g,'&quot;');
      return '<tr style="cursor:pointer" onclick="etfHClickStock(&quot;'+s.c+'&quot;,&quot;'+nn+'&quot;)">'
        +'<td data-order="'+s.c+'"><b style="color:#4fc3f7">'+s.c+'</b> '+s.n+'</td>'
        +'<td class="'+ndvC+'" data-order="'+ndvO+'">'+ndvF+'</td>'
        +'<td class="'+nlC+'" data-order="'+nlO+'">'+(nlF||'-')+'</td>'
        +'<td class="'+sbcC+'" data-order="'+sbcO+'">'+sbcF+'</td>'
        +'<td class="'+chgC+'" data-order="'+chgO+'">'+chgF+'</td>'
        +'<td class="'+vrC+'" data-order="'+vrO+'">'+vrF+'</td>'
        +'<td data-order="'+s.ts+'">'+fS(s.ts)+'</td>'
        +'<td data-order="'+s.tv+'">'+fV(s.tv)+'</td>'
        +'<td class="text-center" data-order="'+s.fc+'">'+s.fc+'</td>'
        +'<td class="'+(s.bv?'pos':'text-muted')+'" data-order="'+bvO+'">'+(s.bv?(s.bv/1e8).toFixed(2)+' 億':'-')+'</td>'
        +'<td class="'+(s.sv?'neg':'text-muted')+'" data-order="'+svO+'">'+(s.sv?'-'+(s.sv/1e8).toFixed(2)+' 億':'-')+'</td>'
        +'</tr>';
    }}).join('');
  }}

  // ── build change table HTML ──
  function _buildChangeRows(changes){{
    return (changes||[]).map(function(x){{
      var tag=x.hp&&x.is_new?'<span class="badge bg-success ms-1">新增</span>'
             :x.hp&&x.is_rm ?'<span class="badge bg-danger ms-1">移除</span>':'';
      var dvF,dvC,dvO;
      if(!x.hp||x.dv===null){{dvF='-';dvC='text-muted';dvO=S0;}}
      else if(x.dv){{dvF=(x.dv>0?'+':'-')+(Math.abs(x.dv)/1e8).toFixed(2)+' 億';dvC=cl(x.dv);dvO=x.dv;}}
      else{{dvF='0.00 億';dvC='text-muted';dvO=0;}}
      var dlF,dlC,dlO;
      if(!x.hp){{dlF='-';dlC='text-muted';dlO=S0;}}
      else if(x.dl){{dlF=(x.dl>0?'+':'')+x.dl.toLocaleString();dlC=cl(x.dl);dlO=x.dl;}}
      else{{dlF='0';dlC='text-muted';dlO=0;}}
      var sbcF,sbcC,sbcO;
      if(!x.sbc&&x.sbc!==0){{sbcF='-';sbcC='text-muted';sbcO=S0;}}
      else{{sbcF=(x.sbc>0?'+':'')+Math.abs(x.sbc).toFixed(3)+'%';sbcC=cl(x.sbc);sbcO=x.sbc;}}
      var chgF=x.chg!=null?(x.chg>0?'+':'')+x.chg.toFixed(2)+'%':'-';
      var chgC=x.chg!=null?cl(x.chg):'text-muted', chgO=x.chg!=null?x.chg:S0;
      return '<tr data-changed="'+(x.changed?1:0)+'" data-etf="'+x.ec+'" data-stock="'+x.c+'">'
        +'<td data-search="'+x.ec+'"><small class="text-muted">'+x.ec+'</small> '+x.en+'</td>'
        +'<td><b style="color:#4fc3f7">'+x.c+'</b> '+x.n+tag+'</td>'
        +'<td data-order="'+x.w+'">'+x.w.toFixed(2)+'%</td>'
        +'<td data-order="'+x.s+'">'+fS(x.s)+'</td>'
        +'<td data-order="'+x.v+'">'+fV(x.v)+'</td>'
        +'<td class="'+dvC+'" data-order="'+dvO+'">'+dvF+'</td>'
        +'<td class="'+dlC+'" data-order="'+dlO+'">'+dlF+'</td>'
        +'<td class="'+sbcC+'" data-order="'+sbcO+'">'+sbcF+'</td>'
        +'<td class="'+chgC+'" data-order="'+chgO+'">'+chgF+'</td>'
        +'</tr>';
    }}).join('');
  }}

  // ── build fund table HTML ──
  function _buildFundRows(etfs){{
    return (etfs||[]).map(function(e){{
      var net=e.nbv-e.sbv,ns=net>0?'+':net<0?'-':'',nc=net<0?'neg':'';
      var nm=ns+(Math.abs(net)/1e8).toFixed(2)+' 億';
      var buys=e.holdings.filter(function(h){{return h.dl>0;}}).sort(function(a,b){{return b.dl-a.dl;}});
      var sells=e.holdings.filter(function(h){{return h.dl<0;}}).sort(function(a,b){{return Math.abs(b.dl)-Math.abs(a.dl);}});
      function _fItem(h){{return '<b style="color:#4fc3f7">'+h.c+'</b> '+(h.dl>0?'+':'')+h.dl.toLocaleString()+'張';}}
      function _trunc(arr){{
        if(!arr.length) return '-';
        var n=arr.length,items=n<=4?arr.map(_fItem):arr.slice(0,3).map(_fItem).concat(['等'+n+'檔']);
        var lines=[];
        for(var i=0;i<items.length;i+=2) lines.push(items.slice(i,i+2).join('<span style="color:#4a5568;margin:0 4px">·</span>'));
        return lines.join('<br>');
      }}
      var chg=buys.length+sells.length, hol=e.holdings.filter(function(h){{return h.s>0;}}).length;
      var prevDate=''; // fund date
      var en2=e.en.replace(/"/g,'&quot;');
      return '<tr style="cursor:pointer" onclick="etfHClickFund(&quot;'+e.ec+'&quot;,&quot;'+en2+'&quot;)">'
        +'<td data-order="'+e.ec+'"><b>'+e.ec+'</b> '+e.en+'</td>'
        +'<td data-order="'+net+'"><span class="'+nc+'">'+nm+'</span><br>'
        +'<small class="fw-normal" style="font-size:.75rem;color:inherit">買+'+(e.nbv/1e8).toFixed(2)+'億 / 賣-'+(e.sbv/1e8).toFixed(2)+'億</small></td>'
        +'<td style="white-space:normal">'+_trunc(buys)+'</td>'
        +'<td class="'+(sells.length?'neg':'')+'" style="white-space:normal">'+_trunc(sells)+'</td>'
        +'<td data-order="'+chg+'"><span style="font-size:1rem;font-weight:700">'+chg+'</span> <small class="text-muted">變動</small><br><small class="text-muted">'+hol+' 持股</small></td>'
        +'<td data-order="'+e.nav+'">'+fV(e.nav)+'</td>'
        +'<td data-order="'+e.date+'">'+e.date+'</td>'
        +'<td></td>'
        +'</tr>';
    }}).join('');
  }}

  // ── render full history view ──
  function _renderFull(data){{
    var hp=data.has_prev, tb=data.total_buy||0, ts_=data.total_sell||0;
    var chgCnt=(data.changes||[]).filter(function(x){{return x.changed;}}).length;
    var totHol=(data.changes||[]).filter(function(x){{return x.s>0;}}).length;
    var sRows=_buildStockRows(data.stocks,hp);
    var cRows=_buildChangeRows(data.changes);
    var fRows=_buildFundRows(data.etfs);
    var actCnt=(data.stocks||[]).filter(function(s){{return s.ts>0;}}).length;
    return '<div class="card mb-3" id="etfHStockCard">'
      +'<div class="card-header px-3 py-2 fw-bold d-flex align-items-center flex-wrap gap-2">'
      +'<span id="etfHStockTitle">股票總表</span>'
      +'<small class="text-muted fw-normal">點股票，看哪些 ETF 持有</small>'
      +'<span id="etfHStockCount" class="fw-normal text-muted" style="font-size:.85rem">共 '+actCnt+' 檔</span>'
      +'<span id="etfHBreadcrumb" class="ms-auto text-muted d-none" style="font-size:.8rem">股票總表 → 持股明細 → 各基金買賣概況</span>'
      +'</div>'
      +_bar(tb,ts_,hp)
      +'<div id="etfHStockFilterBar" class="px-3 pt-2 pb-1 border-bottom d-none">'
      +'<div class="d-flex align-items-center gap-2">'
      +'<input id="etfHStockSearch" class="form-control form-control-sm" style="max-width:160px" placeholder="股票代號" oninput="etfHOnSearch(this.value)">'
      +'<button class="btn btn-sm btn-outline-secondary py-0" onclick="etfHClearStockFilter()">清除篩選</button>'
      +'</div></div>'
      +'<div class="table-responsive"><table class="table table-hover table-sm mb-0" id="etfHStockTable">'
      +'<thead><tr><th>標的</th><th title="淨變動估金額" style="cursor:help">淨額</th>'
      +'<th title="淨變動張數" style="cursor:help">張數</th>'
      +'<th title="淨變動張數／已發行股數" style="cursor:help">股本比</th>'
      +'<th title="當日漲跌幅" style="cursor:help">漲跌%</th>'
      +'<th title="淨變動張數／當日成交量" style="cursor:help">量比</th>'
      +'<th>持有張</th><th>持有估值</th>'
      +'<th title="持有本股的基金數量" style="cursor:help">基金</th>'
      +'<th title="買進估金額" style="cursor:help">買進</th>'
      +'<th title="賣出估金額" style="cursor:help">賣出</th>'
      +'</tr></thead><tbody>'+sRows+'</tbody></table></div></div>'

      +'<div class="card mb-3" id="etfHChangeCard">'
      +'<div class="card-header px-3 py-2 fw-bold d-flex align-items-center flex-wrap gap-2">'
      +'<span id="etfHChangeTitle">本日變動持股明細</span>'
      +'<span id="etfHChangeCount" class="fw-normal text-muted" style="font-size:.85rem">變動 '+chgCnt+' 筆 / 全部持股 '+totHol+' 筆</span>'
      +'<span id="etfHFundChip" class="ms-1"></span>'
      +'<span id="etfHStockChip" class="ms-1"></span>'
      +'<div class="ms-auto d-flex gap-2">'
      +'<button id="etfHClearFundBtn" class="btn btn-sm btn-outline-secondary py-0 d-none" onclick="etfHClearFund()">清除基金</button>'
      +'<button id="etfHClearStockBtn" class="btn btn-sm btn-outline-secondary py-0 d-none" onclick="etfHClearStock()">清除股票</button>'
      +'</div></div>'
      +'<div class="px-3 pt-2 pb-1 border-bottom d-flex align-items-center gap-3">'
      +'<button id="etfHChangedOnlyBtn" class="btn btn-sm btn-primary py-0 px-2" onclick="etfHToggleChanged()">本日有加減碼</button>'
      +'<small id="etfHSortHint" class="text-muted">依變動估金額降冪排序；共 '+chgCnt+' 筆變動</small>'
      +'</div>'
      +'<div class="table-responsive"><table class="table table-hover table-sm mb-0" id="etfHChangeTable">'
      +'<thead><tr><th>基金</th><th>股票</th><th>比例</th><th>張數</th><th>估市值</th>'
      +'<th title="持股變動對應的估算金額" style="cursor:help">變動估金額</th>'
      +'<th title="變動張數" style="cursor:help">△張</th>'
      +'<th title="變動張數／已發行股數" style="cursor:help">△股本比</th>'
      +'<th title="當日漲跌幅" style="cursor:help">日漲跌</th>'
      +'</tr></thead><tbody>'+cRows+'</tbody></table></div></div>'

      +'<div class="card mb-3" id="etfHFundCard">'
      +'<div class="card-header px-3 py-2 fw-bold">各基金買賣概況　'+(data.etfs||[]).length+' 檔</div>'
      +'<div class="table-responsive"><table class="table table-hover table-sm mb-0" id="etfHFundTable">'
      +'<thead><tr><th>基金</th><th>淨買賣</th><th>加碼</th><th>減碼</th><th>變動</th><th>規模</th><th>日期</th><th></th>'
      +'</tr></thead><tbody>'+fRows+'</tbody></table></div></div>';
  }}

  // ── init all history DataTables + custom filter ──
  function _initAllDTs(){{
    _destroyHDTs();
    _hChangedOnly=true; _hFundFilter=null; _hStockFilter=null;

    // change table: custom filter for changedOnly + fund/stock
    if($.fn.DataTable){{
      $.fn.dataTable.ext.search=$.fn.dataTable.ext.search.filter(function(f){{
        return f._etfH!==true;
      }});
      var hFilter=function(s,d,idx,row){{
        if(!s.nTable||s.nTable.id!=='etfHChangeTable') return true;
        var nTr=s.aoData&&s.aoData[idx]?s.aoData[idx].nTr:null;
        var $tr=nTr?$(nTr):null;
        if(_hFundFilter && (!$tr||$tr.attr('data-etf')!==_hFundFilter)) return false;
        if(_hStockFilter && (!$tr||$tr.attr('data-stock')!==_hStockFilter)) return false;
        if(_hChangedOnly && !_hFundFilter && !_hStockFilter && (!$tr||$tr.attr('data-changed')!=='1')) return false;
        return true;
      }};
      hFilter._etfH=true;
      $.fn.dataTable.ext.search.push(hFilter);
    }}

    _initDT('etfHStockTable',1,[
      {{targets:[1,2,3,4,5,6,7,8,9,10],type:'num'}},
      {{targets:0,width:'110px'}},
    ]);
    _initDT('etfHChangeTable',5,[
      {{targets:[2,3,4,5,6,7,8],type:'num'}},
    ]);
    _initDT('etfHFundTable',1,[
      {{targets:[1,4,5,6],type:'num'}},
    ]);
  }}

  // ── history filter functions ──
  window.etfHClickStock=function(code,name){{
    _hStockFilter=code; _hFundFilter=null;
    if(_hDTs['etfHStockTable']) _hDTs['etfHStockTable'].search(code).draw();
    $('#etfHStockFilterBar').removeClass('d-none');
    $('#etfHStockSearch').val(code);
    $('#etfHBreadcrumb').removeClass('d-none');
    if(_hDTs['etfHChangeTable']){{
      _hDTs['etfHChangeTable'].order([[3,'desc']]).draw();
    }}
    $('#etfHChangeTitle').text(code+' '+name+'：主動 ETF 持股');
    $('#etfHStockChip').html('<span class="badge bg-success">'+code+' '+name+'</span>');
    $('#etfHFundChip').html('');
    $('#etfHClearStockBtn').removeClass('d-none');
    $('#etfHClearFundBtn').addClass('d-none');
    $('#etfHChangedOnlyBtn').addClass('d-none');
    $('#etfHSortHint').text('依張數降冪排序');
    var $card=$('#etfHChangeCard');
    if($card.length) $card[0].scrollIntoView({{behavior:'smooth',block:'start'}});
  }};

  window.etfHClickFund=function(ec,en){{
    _hFundFilter=ec; _hStockFilter=null;
    if(_hDTs['etfHChangeTable']) _hDTs['etfHChangeTable'].order([[2,'desc']]).draw();
    $('#etfHChangeTitle').text('持股明細');
    var vis=_hDTs['etfHChangeTable']?_hDTs['etfHChangeTable'].rows({{search:'applied'}}).count():0;
    var tot=_hDTs['etfHChangeTable']?_hDTs['etfHChangeTable'].rows().count():0;
    $('#etfHChangeCount').html('<span class="text-primary fw-bold">'+ec+' '+en+'</span> 符合篩選 '+vis+' 筆 / 全部 '+tot+' 筆');
    $('#etfHFundChip').html('<span class="badge bg-primary">'+ec+' '+en+'</span>');
    $('#etfHClearFundBtn').removeClass('d-none');
    $('#etfHChangedOnlyBtn').addClass('d-none');
    $('#etfHSortHint').text('依比例降冪排序；'+en+' 全部持股');
    var $card=$('#etfHChangeCard');
    if($card.length) $card[0].scrollIntoView({{behavior:'smooth',block:'start'}});
  }};

  window.etfHClearFund=function(){{
    _hFundFilter=null;
    if(_hDTs['etfHChangeTable']) _hDTs['etfHChangeTable'].order([[5,'desc']]).draw();
    var chg=0,tot=_hDTs['etfHChangeTable']?_hDTs['etfHChangeTable'].rows().count():0;
    if(_hDTs['etfHChangeTable']) _hDTs['etfHChangeTable'].rows().every(function(){{if($(this.node()).attr('data-changed')==='1') chg++;}});
    $('#etfHChangeTitle').text('本日變動持股明細');
    $('#etfHChangeCount').text('變動 '+chg+' 筆 / 全部持股 '+tot+' 筆');
    $('#etfHFundChip').html('');
    $('#etfHClearFundBtn').addClass('d-none');
    $('#etfHChangedOnlyBtn').removeClass('d-none');
    $('#etfHSortHint').text('依變動估金額降冪排序；共 '+chg+' 筆變動');
  }};

  window.etfHClearStock=function(){{
    _hStockFilter=null;
    etfHClearStockFilter();
    if(_hDTs['etfHChangeTable']) _hDTs['etfHChangeTable'].order([[5,'desc']]).draw();
    var chg=0,tot=_hDTs['etfHChangeTable']?_hDTs['etfHChangeTable'].rows().count():0;
    if(_hDTs['etfHChangeTable']) _hDTs['etfHChangeTable'].rows().every(function(){{if($(this.node()).attr('data-changed')==='1') chg++;}});
    $('#etfHChangeTitle').text('本日變動持股明細');
    $('#etfHChangeCount').text('變動 '+chg+' 筆 / 全部持股 '+tot+' 筆');
    $('#etfHStockChip').html('');
    $('#etfHClearStockBtn').addClass('d-none');
    $('#etfHChangedOnlyBtn').removeClass('d-none');
    $('#etfHSortHint').text('依變動估金額降冪排序；共 '+chg+' 筆變動');
  }};

  window.etfHOnSearch=function(val){{
    if(!_hDTs['etfHStockTable']) return;
    _hDTs['etfHStockTable'].search(val).draw();
    var m=_hDTs['etfHStockTable'].rows({{search:'applied'}}).count();
    var t=_hDTs['etfHStockTable'].rows().count();
    $('#etfHStockCount').text('符合篩選 '+m+' 檔 / 全部 '+t+' 檔');
  }};

  window.etfHClearStockFilter=function(){{
    if(_hDTs['etfHStockTable']){{_hDTs['etfHStockTable'].search('').draw();}}
    $('#etfHStockFilterBar').addClass('d-none');
    $('#etfHStockSearch').val('');
    $('#etfHBreadcrumb').addClass('d-none');
    var t=_hDTs['etfHStockTable']?_hDTs['etfHStockTable'].rows().count():'';
    $('#etfHStockCount').text('共 '+t+' 檔');
  }};

  window.etfHToggleChanged=function(){{
    _hChangedOnly=!_hChangedOnly;
    var $b=$('#etfHChangedOnlyBtn');
    $b.toggleClass('btn-primary',_hChangedOnly).toggleClass('btn-outline-primary',!_hChangedOnly);
    if(_hDTs['etfHChangeTable']) _hDTs['etfHChangeTable'].draw();
  }};

  window.etfSelectDate=function(key){{
    var sel=document.getElementById('etfDateSelect');
    if(sel && sel.value!==key) sel.value=key;
    var isLive=(key===LIVE_KEY);
    var _cv=document.getElementById('etfCurrentView');
    if(_cv) _cv.style.display=isLive?'':'none';
    var hv=document.getElementById('etfHistoryView');
    if(isLive){{
      _destroyHDTs();
      if(typeof $!=='undefined'&&$.fn&&$.fn.dataTable)
        $.fn.dataTable.ext.search=$.fn.dataTable.ext.search.filter(function(f){{return f._etfH!==true;}});
      hv.style.display='none'; hv.innerHTML='';
      return;
    }}
    hv.style.display='';
    hv.innerHTML='<div class="text-muted p-3">載入中…</div>';
    fetch('etf_data/summary_'+key+'.json?_='+Date.now())
      .then(function(r){{if(!r.ok)throw new Error(r.status);return r.json();}})
      .then(function(data){{
        _hData=data;
        var label=key.slice(0,3)+'/'+key.slice(3,5)+'/'+key.slice(5,7);
        hv.innerHTML='<div class="alert alert-secondary py-1 px-3 mb-2" style="font-size:.85rem">歷史紀錄：'+label+'</div>'+_renderFull(data);
        setTimeout(function(){{_initAllDTs();}},0);
      }})
      .catch(function(){{hv.innerHTML='<div class="text-danger p-3">載入失敗，此日期無資料</div>';}});
  }};

  // 預設顯示最新（第一項）
  if(AVAIL.length>0) etfSelectDate(AVAIL[0].key);
}})();
</script>"""


if __name__ == "__main__":
    print("【主動ETF追蹤】")
    results = run_all()
    if results:
        html = generate_etf_html(results)
        out = Path(__file__).parent / "etf_test.html"
        out.write_text(f"<html><head><meta charset='utf-8'><link rel='stylesheet' href='https://cdn.jsdelivr.net/npm/bootstrap@5/dist/css/bootstrap.min.css'></head><body>{html}</body></html>", encoding="utf-8")
        print(f"預覽：{out}")
